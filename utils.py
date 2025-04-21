import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from fuzzywuzzy import fuzz
import io
import re
import os

# Глобальные переменные
# Колонки, которые не должны переноситься при копировании данных
excluded_columns = ['Артикул WB', 'Название модели (для объединения в одну карточку)*']

def load_excel_file(file):
    """
    Загружает Excel файл и возвращает объект рабочей книги и список листов.
    Обрабатывает и преобразует специфические для маркетплейсов поля (артикулы, SKU) в строковый формат
    для избежания ошибок конвертации.
    
    Args:
        file: Загруженный файл в формате BytesIO
        
    Returns:
        Tuple: (workbook, list_of_sheets)
    """
    try:
        # Загружаем с data_only=True, чтобы получить значения, а не формулы
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheets = workbook.sheetnames
        
        # Обрабатываем все листы для преобразования идентификаторов в строки
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            
            # Находим заголовки, которые содержат идентификаторы (артикулы, SKU и т.д.)
            id_columns = []  # Столбцы, содержащие идентификаторы
            header_row = 1  # По умолчанию проверяем первые 5 строк
            
            # Ищем заголовки в первых 5 строках
            for row_idx in range(1, 6):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        cell_value_lower = cell.value.lower()
                        # Проверяем, содержит ли заголовок ключевые слова для идентификаторов
                        if any(id_word in cell_value_lower for id_word in ['артикул', 'sku', 'guid', 'штрихкод', 'баркод', 'код']):
                            id_columns.append(col_idx)
                            header_row = row_idx
            
            # Если нашли идентификаторы, преобразуем их значения в строки
            if id_columns:
                for col_idx in id_columns:
                    for row_idx in range(header_row + 1, sheet.max_row + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            # Преобразуем все в строку, включая числа
                            cell.value = str(cell.value)
        
        return workbook, sheets
    except Exception as e:
        error_str = str(e)
        if "expected <class 'openpyxl.worksheet.cell_range.MultiCellRange'>" in error_str:
            # Специальная обработка для ошибки шаблонов Wildberries
            raise Exception("⚠️ Пожалуйста, пересохраните файл перед загрузкой. Файлы шаблонов Wildberries могут содержать специальные форматы, требующие пересохранения.")
        else:
            raise Exception(f"Ошибка при загрузке Excel файла: {error_str}")
            
def find_header_row(worksheet, sheet_name=None, max_rows=30):
    """
    Находит строку заголовков в Excel файле с учетом типичного расположения для каждого маркетплейса.
    Использует точные данные о расположении заголовков:
    - Ozon: строка 2 на листе "Шаблон"
    - Wildberries: строка 3 на листе "Товары"
    - ЛеманПро: строка 4 (имя листа часто совпадает с именем файла)
    - Яндекс.Маркет: строка 2 на листе "Данные о товарах"
    
    Args:
        worksheet: Лист Excel
        sheet_name: Имя листа (для специальной обработки шаблонов)
        max_rows: Максимальное количество строк для поиска заголовков
        
    Returns:
        int: Номер строки с заголовками (начиная с 1) или 1, если не найдено
    """
    if not sheet_name:
        sheet_name = ""
    
    sheet_name_lower = sheet_name.lower()
    
    # ПРАВИЛО 1: Ozon - строка 2 на листе "Шаблон"
    if "шаблон" in sheet_name_lower or "template" in sheet_name_lower or "ozon" in sheet_name_lower:
        # Проверим заголовки второй строки на типичные для Ozon
        if worksheet.max_row >= 2:
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[2]]
            if any('артикул*' in val for val in row_values) or any('название товара*' in val for val in row_values):
                return 2
    
    # ПРАВИЛО 2: Wildberries - строка 3 на листе "Товары"
    if "товары" in sheet_name_lower or "wildberries" in sheet_name_lower or "wb" in sheet_name_lower:
        # Проверим заголовки третьей строки на типичные для WB
        if worksheet.max_row >= 3:
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[3]]
            if any('артикул продавца' in val for val in row_values) or any('артикул wb' in val for val in row_values):
                return 3
    
    # ПРАВИЛО 3: Яндекс.Маркет - проверяем и строку 2, и строку 4 на листе "Данные о товарах"
    if "данные о товарах" in sheet_name_lower or "яндекс" in sheet_name_lower or "market" in sheet_name_lower:
        # Сначала проверим заголовки в 4-й строке (приоритетно)
        if worksheet.max_row >= 4:
            row_values_4 = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[4]]
            if any('ваш sku' in val for val in row_values_4) or any('качество карточки' in val for val in row_values_4):
                return 4
                
        # Затем проверим заголовки во 2-й строке
        if worksheet.max_row >= 2:
            row_values_2 = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[2]]
            if any('ваш sku' in val for val in row_values_2) or any('качество карточки' in val for val in row_values_2):
                return 2
    
    # ПРАВИЛО 4: ЛеманПро - строка 4, имя листа часто совпадает с именем файла
    # Для ЛеманПро сначала проверим 4-ю строку
    if "леман" in sheet_name_lower or "leman" in sheet_name_lower or "атем" in sheet_name_lower or "atem" in sheet_name_lower:
        if worksheet.max_row >= 4:
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[4]]
            if any('guid' in val for val in row_values) or any('товара мерчанта' in val for val in row_values):
                return 4
    
    # Если не сработали правила по имени листа, проверяем содержимое по строкам
    
    # Сначала проверяем типичные строки заголовков для каждого маркетплейса
    header_rows_to_check = [
        (2, ['артикул*', 'название товара*', 'цена, руб.*', 'ozon id']),  # Ozon
        (3, ['артикул продавца', 'артикул wb', 'наименование', 'фото']),  # Wildberries
        (4, ['guid', 'наименование товара мерчанта', 'бренд товара']),     # ЛеманПро
        (4, ['ваш sku', 'ваш sku *', 'название товара *', 'качество карточки']), # Яндекс.Маркет (строка 4)
        (2, ['ваш sku', 'ваш sku *', 'название товара *', 'качество карточки'])  # Яндекс.Маркет (строка 2)
    ]
    
    for row_idx, markers in header_rows_to_check:
        if worksheet.max_row >= row_idx:
            row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[row_idx]]
            matches = sum(1 for marker in markers if any(marker in val for val in row_values))
            if matches >= 2:  # Если нашли хотя бы 2 маркера
                return row_idx
    
    # Если не нашли по типичным строкам, проверяем все строки до max_rows
    # Определим характерные маркеры для всех маркетплейсов
    all_markers = [
        'артикул', 'название', 'наименование', 'цена', 'бренд', 'фото', 'sku',
        'guid', 'баркод', 'штрихкод', 'группа', 'категория'
    ]
    
    for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
        row_values = [str(cell.value).strip().lower() if cell.value else '' for cell in worksheet[row_idx]]
        
        # Считаем количество найденных маркеров
        matches = sum(1 for marker in all_markers if any(marker in val for val in row_values))
        
        # Также проверяем количество непустых ячеек
        non_empty_count = sum(1 for val in row_values if val)
        
        # Если нашли много маркеров или много непустых ячеек, считаем это строкой заголовков
        if matches >= 3 or non_empty_count >= 5:
            return row_idx
    
    # Если ничего не нашли, возвращаем 1 (первая строка)
    return 1

def save_excel_file(workbook):
    """
    Сохраняет Excel файл и возвращает BytesIO объект
    
    Args:
        workbook: Объект рабочей книги openpyxl
        
    Returns:
        BytesIO: Объект байтового потока для скачивания
    """
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def find_best_marketplace_sheet(workbook):
    """
    Ищет в книге Excel лист, наиболее подходящий для маркетплейса (Ozon, Wildberries, ЛеманПро или Яндекс.Маркет)
    на основе анализа структуры листа, позиции заголовков и характерных полей.
    
    Args:
        workbook: Объект рабочей книги openpyxl
        
    Returns:
        tuple: (имя_листа, тип_маркетплейса, строка_с_заголовками)
    """
    # 1. Жесткая обработка для шаблона АТЁМ - важно, чтобы работало во всех случаях
    for sheet_name in workbook.sheetnames:
        if "атём" in sheet_name.lower() or "атем" in sheet_name.lower() or "atem" in sheet_name.lower():
            return sheet_name, 'lemanpro', 4
    
    # Характерные признаки для каждого маркетплейса на основе реальных примеров заголовков
    marketplace_signatures = {
        'ozon': {
            'unique_fields': [
                'артикул*', 'название товара*', 'название товара', 'бренд*', 'бренд', 'цена, руб.*', 'цена, руб',
                'ндс, %*', 'вес в упаковке, г*', 'ширина упаковки, мм*', 'длина упаковки, мм*', 
                'высота упаковки, мм*', 'ссылка на главное фото*', 'ссылки на дополнительные фото',
                'ozon id', 'название модели', 'тип*'
            ],
            'common_sheet_names': ['шаблон', 'template', 'import', 'товары ozon', 'озон'],
            'header_rows': [2, 1],
            'units': ['мм', 'г'],  # Использует миллиметры и граммы
            'asterisk': True  # Обязательные поля помечены *
        },
        'wildberries': {
            'unique_fields': [
                'артикул wb', 'артикул продавца', 'категория продавца', 'баркод', 'штрихкод',
                'группа', 'наименование', 'бренд', 'описание', 'фото', 'видео', 'цвет', 
                'вес с упаковкой (кг)', 'ставка ндс', 'высота упаковки', 'ширина упаковки',
                'длина упаковки'
            ],
            'common_sheet_names': ['товары', 'products', 'карточки', 'номенклатуры', 'wildberries', 'вб'],
            'header_rows': [3, 2, 1],
            'units': ['кг'],  # Использует килограммы
            'asterisk': False  # Не использует звездочки для обязательных полей
        },
        'lemanpro': {
            'unique_fields': [
                'guid', 'наименование товара мерчанта', 'бренд товара', 'модель товара', 
                'артикул товара', 'серия/коллекция', 'штрих-код', 'размеры в упаковке: ширина (мм)',
                'размеры в упаковке: длина (мм)', 'размеры в упаковке: высота (мм)', 
                'цветовая палитра', 'тип упаковки', 'тип продукта', 'основной материал',
                'нетто', 'страна производства'
            ],
            'common_sheet_names': ['леманпро', 'leman', 'lp', 'товары леман'],
            'header_rows': [4, 3],
            'units': ['мм', 'кг'],  # Использует миллиметры и килограммы
            'asterisk': False,  # Не использует звездочки для обязательных полей
            'row4_headers': ['guid', 'код тн вэд', 'наименование товара мерчанта', 'бренд товара', 'модель товара'] # Типичные первые 5 заголовков ЛеманПро в строке 4
        },
        'yandex': {
            'unique_fields': [
                'ваш sku', 'ваш sku *', 'качество карточки', 'рекомендации по заполнению', 
                'название группы вариантов', 'название товара *', 'название товара', 
                'ссылка на изображение *', 'ссылка на изображение', 'изображение для миниатюры', 
                'бренд *', 'бренд', 'штрихкод *', 'штрихкод', 'теги', 'габариты с упаковкой, см',
                'цена *', 'зачёркнутая цена', 'sku на маркете', 'в архиве',
                'грузоподъемность, кг', 'диаметр колеса, см'
            ],
            'common_sheet_names': ['данные о товарах', 'товары яндекс', 'яндекс маркет', 'yandex', 'маркет'],
            'header_rows': [4, 2, 1],  # Добавляем строку 4 как приоритетную для Яндекс.Маркет
            'units': ['см', 'кг'],  # Использует сантиметры и килограммы
            'asterisk': True,  # Обязательные поля помечены *
            'row4_headers': ['ваш sku *', 'качество карточки', 'рекомендации по заполнению', 'название товара *', 'ссылка на изображение *'] # Типичные первые 5 заголовков Яндекс.Маркет в строке 4
        }
    }
    
    # Словарь для маппинга уникальных идентификаторов товара между площадками
    id_mappings = {
        'guid': 'lemanpro',
        'артикул товара': 'lemanpro',
        'артикул продавца': 'wildberries',
        'ваш sku': 'yandex',
        'ваш sku *': 'yandex',
        'артикул*': 'ozon',
        'артикул wb': 'wildberries'
    }
    
    # Функция для нормализации текста
    def normalize_text(text):
        if text is None:
            return ''
        return str(text).lower().strip()
    
    # 2. Проверяем характерные имена листов
    for sheet_name in workbook.sheetnames:
        sheet_name_lower = sheet_name.lower()
        
        for marketplace, signature in marketplace_signatures.items():
            for common_name in signature['common_sheet_names']:
                if common_name in sheet_name_lower:
                    # Нашли потенциальное совпадение по имени листа
                    sheet = workbook[sheet_name]
                    
                    # Проверяем характерные строки заголовков для данного маркетплейса
                    for header_row in signature['header_rows']:
                        if sheet.max_row >= header_row:
                            header_values = [normalize_text(cell.value) for cell in sheet[header_row]]
                            
                            # Проверяем наличие характерных полей
                            unique_fields_found = sum(1 for field in signature['unique_fields'] 
                                                     if any(field in val for val in header_values))
                            
                            # Если нашли достаточно характерных полей, считаем, что определили маркетплейс
                            if unique_fields_found >= 3:  # Минимум 3 характерных поля для надежного определения
                                return sheet_name, marketplace, header_row
    
    # 3. Если по имени листа не определили, проходим по всем листам и ищем характерные заголовки
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Проверяем строки с 1 по 5, где могут быть заголовки
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            if sheet.max_row >= row_idx:
                row_values = [normalize_text(cell.value) for cell in sheet[row_idx]]
                
                # Проверяем каждый маркетплейс
                marketplace_scores = {}
                
                for marketplace, signature in marketplace_signatures.items():
                    score = 0
                    
                    # Проверяем наличие уникальных полей
                    for field in signature['unique_fields']:
                        for val in row_values:
                            if field in val:
                                score += 2  # Более высокий вес для уникальных полей
                                break  # Идем к следующему полю
                    
                    # Проверяем наличие звездочек (характерно для Ozon и Яндекс.Маркет)
                    has_asterisks = any('*' in val for val in row_values if val)
                    if has_asterisks == signature['asterisk']:
                        score += 3
                    
                    # Проверяем используемые единицы измерения
                    for unit in signature['units']:
                        if any(unit in val for val in row_values):
                            score += 1
                    
                    # Проверяем идентификаторы товаров 
                    for id_field, id_marketplace in id_mappings.items():
                        if id_marketplace == marketplace and any(id_field in val for val in row_values):
                            score += 4  # Высокий вес для полей-идентификаторов
                    
                    # Особая проверка для ЛеманПро - заголовки в 4-й строке
                    if row_idx == 4 and marketplace == 'lemanpro':
                        score += 5
                        
                    # Особая проверка для Ozon - заголовки во 2-й строке
                    if row_idx == 2 and marketplace == 'ozon' and has_asterisks:
                        score += 5
                        
                    # Особая проверка для WB - заголовки в 3-й строке
                    if row_idx == 3 and marketplace == 'wildberries' and any('артикул wb' in val for val in row_values):
                        score += 5
                    
                    # Сохраняем оценку для маркетплейса
                    marketplace_scores[marketplace] = score
                
                # Если есть хотя бы один маркетплейс с высоким рейтингом, возвращаем его
                best_marketplace = max(marketplace_scores.items(), key=lambda x: x[1])
                if best_marketplace[1] >= 5:  # Минимальный порог оценки для надежного определения
                    return sheet_name, best_marketplace[0], row_idx
    
    # 4. Проверка для шаблонов Ozon (характерная черта - лист "Шаблон")
    if 'Шаблон' in workbook.sheetnames:
        sheet = workbook['Шаблон']
        # В шаблонах Ozon заголовки обычно на второй строке
        if sheet.max_row >= 2:
            row_values = [str(cell.value).strip() if cell.value else '' for cell in sheet[2]]
            header_indicators = ['артикул', 'название', 'фото', 'бренд', 'цена']
            found_headers = 0
            for val in row_values:
                if any(indicator in val.lower() for indicator in header_indicators):
                    found_headers += 1
                    
            if found_headers >= 1:
                return 'Шаблон', 'ozon', 2
    
    # 5. Если всё еще не нашли, применяем эвристику - ищем строку с большим количеством ячеек
    max_cells = 0
    max_cells_row = 1
    max_cells_sheet = workbook.sheetnames[0]
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_idx in range(1, min(15, sheet.max_row + 1)):
            non_empty = sum(1 for cell in sheet[row_idx] if cell.value)
            if non_empty > max_cells:
                max_cells = non_empty
                max_cells_row = row_idx
                max_cells_sheet = sheet_name
    
    # Если нашли строку с хотя бы 5 ячейками, считаем её заголовком
    if max_cells >= 5:
        # Попытаемся определить тип маркетплейса по содержимому заголовков
        header_values = [normalize_text(cell.value) for cell in workbook[max_cells_sheet][max_cells_row]]
        
        # Проверяем специфические идентификаторы для каждого маркетплейса
        for id_field, marketplace in id_mappings.items():
            if any(id_field in val for val in header_values):
                return max_cells_sheet, marketplace, max_cells_row
        
        # Проверяем наличие звездочек
        has_asterisks = any('*' in val for val in header_values if val)
        if has_asterisks:
            # Если есть звездочки, скорее всего это Ozon или Яндекс
            if any('мм' in val for val in header_values):
                return max_cells_sheet, 'ozon', max_cells_row
            elif any('см' in val for val in header_values):
                return max_cells_sheet, 'yandex', max_cells_row
        else:
            # Если нет звездочек, проверяем другие признаки
            if any('артикул wb' in val for val in header_values):
                return max_cells_sheet, 'wildberries', max_cells_row
            elif any('артикул продавца' in val for val in header_values):
                return max_cells_sheet, 'wildberries', max_cells_row
            elif any('guid' in val for val in header_values) or max_cells_row == 4:
                return max_cells_sheet, 'lemanpro', max_cells_row
        
        return max_cells_sheet, 'other', max_cells_row
    
    # Если ничего не нашли, возвращаем первый лист и строку 1
    return workbook.sheetnames[0], 'other', 1

def detect_marketplace_template(columns):
    """
    Определяет, к какому маркетплейсу относится таблица по её заголовкам,
    используя простую проверку на наличие характерных заголовков маркетплейса.
    
    Args:
        columns: Список названий колонок
        
    Returns:
        tuple: (маркетплейс ('ozon', 'wildberries', 'lemanpro', 'yandex', 'other'), уровень уверенности)
    """
    if not columns:
        return "other", 0
    
    # Нормализуем заголовки для простоты поиска
    normalized_columns = [str(col).lower().strip() if col is not None else '' for col in columns]
    
    # Проверка для OZON
    ozon_key_fields = ['название товара*', 'артикул*', 'цена, руб.*', 'ндс, %*', 'бренд*', 'обязательное поле']
    ozon_matches = 0
    for field in ozon_key_fields:
        if any(field in col for col in normalized_columns):
            ozon_matches += 1
    
    # Если нашли достаточно полей Ozon
    if ozon_matches >= 3 or any('обязательное поле' in col for col in normalized_columns):
        return 'ozon', 95.0
    
    # Проверка для Wildberries
    wb_key_fields = ['артикул продавца', 'артикул wb', 'наименование', 'группа', 'фото']
    wb_matches = 0
    for field in wb_key_fields:
        if any(field in col for col in normalized_columns):
            wb_matches += 1
    
    # Если нашли достаточно полей Wildberries
    if wb_matches >= 3 or any('артикул wb' in col for col in normalized_columns):
        return 'wildberries', 95.0
    
    # Проверка для ЛеманПро
    lemanpro_key_fields = ['guid', 'код тн вэд', 'наименование товара мерчанта', 'бренд товара', 'модель товара']
    lemanpro_matches = 0
    for field in lemanpro_key_fields:
        if any(field in col for col in normalized_columns):
            lemanpro_matches += 1
    
    # Если нашли достаточно полей ЛеманПро или есть ключевое поле GUID
    if lemanpro_matches >= 3 or any('guid' in col for col in normalized_columns):
        return 'lemanpro', 95.0
    
    # Проверка для Яндекс.Маркет
    yandex_key_fields = ['ваш sku *', 'качество карточки', 'рекомендации по заполнению', 
                         'название группы вариантов', 'название товара *', 'ссылка на изображение *']
    
    # Добавим дополнительные поля, которые встречаются в разных вариантах шаблонов Яндекс.Маркет
    yandex_additional_fields = ['ваш sku', 'уникальный идентификатор товара', 'входит в категорию', 
                               'не входит в категорию', 'param_names', 'param_ids', 'header']
    
    yandex_matches = 0
    
    # Основная проверка по ключевым полям
    for field in yandex_key_fields:
        if any(field in col for col in normalized_columns):
            yandex_matches += 1
    
    # Дополнительная проверка по характерным для Яндекс.Маркет полям
    for field in yandex_additional_fields:
        if any(field == col for col in normalized_columns):  # Точное совпадение
            yandex_matches += 1
    
    # Проверка на особые маркеры в Яндекс шаблоне (header, param_names, param_ids)
    special_markers = sum(1 for col in normalized_columns if col in ['header', 'param_names', 'param_ids'])
    if special_markers >= 2:
        yandex_matches += 3  # Большой вес для этих специальных маркеров
    
    # Если нашли достаточно полей Яндекс или есть характерное поле "Ваш SKU" или специальные маркеры
    if yandex_matches >= 3 or any('ваш sku' in col for col in normalized_columns) or special_markers >= 2:
        return 'yandex', 95.0
    
    # Если не удалось определить по ключевым полям, пробуем по общему числу совпадений
    matches = {
        'ozon': ozon_matches,
        'wildberries': wb_matches,
        'lemanpro': lemanpro_matches,
        'yandex': yandex_matches
    }
    
    best_marketplace, max_count = max(matches.items(), key=lambda x: x[1])
    
    if max_count >= 2:
        confidence = min(80.0, max_count * 20)
        return best_marketplace, confidence
    
    # Если не нашли достаточно совпадений
    return 'other', 0

def map_columns_automatically(source_columns, target_columns, threshold=70):
    """
    Автоматически сопоставляет колонки на основе схожести названий
    
    Args:
        source_columns: Список колонок исходной таблицы
        target_columns: Список колонок целевой таблицы
        threshold: Порог схожести для сопоставления (0-100)
        
    Returns:
        Dict: Словарь соответствия {source_column: target_column}
    """
    # Готовые словари соответствия между маркетплейсами на основе реальных данных
    # Словарь WB -> Ozon
    wb_to_ozon = {
        'Артикул продавца': 'Артикул*',
        'Наименование': 'Название товара*',
        'Бренд': 'Бренд*',
        'Описание': 'Аннотация',
        'Фото': ['Ссылка на главное фото*', 'Ссылки на дополнительные фото'],  # Мапим на два поля в Ozon
        'Вес с упаковкой (кг)': 'Вес в упаковке, г*',
        'Цвет': 'Цвет товара*',
        'Цена': 'Цена, руб.*',
        'Вес без упаковки (кг)': 'Вес товара, г',
        'Высота упаковки': 'Высота упаковки, мм*',
        'Длина упаковки': 'Длина упаковки, мм*',
        'Ширина упаковки': 'Ширина упаковки, мм*',
        'Комплектация': 'Комплектация',
        'Страна производства': 'Страна-изготовитель',
        'Материал изделия': 'Материал',
        'Баркод': 'Штрихкод (Серийный номер / EAN)',
        'Баркоды': 'Штрихкод (Серийный номер / EAN)',
        'Штрихкод': 'Штрихкод (Серийный номер / EAN)',
        'ТНВЭД': 'ТН ВЭД коды ЕАЭС',
        'Количество колес': 'Количество колес тачки',
        'Диаметр колес': 'Диаметр колеса, мм',
        'Грузоподъемность': 'Макс. нагрузка, кг',
        'Объем': 'Объем, л',
        'Ставка НДС': 'НДС, %*',
        'Пол': 'Хештеги'  # Может использоваться для хештегов типа #мужской #женский
        # 'Категория продавца': не маппится, используется имя файла
    }
    
    # Словарь Ozon -> WB
    ozon_to_wb = {
        'Артикул*': 'Артикул продавца',
        'Название товара*': 'Наименование',
        'Бренд*': 'Бренд',
        'Аннотация': 'Описание',
        'Ссылка на главное фото*': 'Фото',
        'Ссылки на дополнительные фото': 'Фото',  # Оба поля с фото в Ozon мапятся на одно в WB
        'Вес в упаковке, г*': 'Вес с упаковкой (кг)',
        'Цвет товара*': 'Цвет',
        'Цена, руб.*': 'Цена',
        'Вес товара, г': 'Вес без упаковки (кг)',
        'Высота упаковки, мм*': 'Высота упаковки',
        'Длина упаковки, мм*': 'Длина упаковки',
        'Ширина упаковки, мм*': 'Ширина упаковки',
        'Комплектация': 'Комплектация',
        'Страна-изготовитель': 'Страна производства',
        'Материал': 'Материал изделия',
        'Штрихкод (Серийный номер / EAN)': ['Баркод', 'Баркоды', 'Штрихкод'],
        'ТН ВЭД коды ЕАЭС': 'ТНВЭД',
        'Количество колес тачки': 'Количество колес',
        'Диаметр колеса, мм': 'Диаметр колес',
        'Макс. нагрузка, кг': 'Грузоподъемность',
        'Объем, л': 'Объем',
        'НДС, %*': 'Ставка НДС',
        'Хештеги': 'Пол'  # Может использоваться для хештегов типа #мужской #женский
        # 'Категория продавца': Не маппится, для WB шаблона используется имя файла
    }
    
    # Словарь WB -> Яндекс
    wb_to_yandex = {
        'Артикул продавца': 'Ваш SKU *',
        'Наименование': 'Название товара *',
        'Бренд': 'Бренд *',
        'Описание': 'Описание товара *',
        'Фото': 'Ссылка на изображение *',
        'Видео': 'Ссылка на видео',
        'Вес с упаковкой (кг)': 'Вес с упаковкой, кг',
        'Цвет': ['Название цвета от производителя', 'Цвет для фильтра'],
        'Цена': 'Цена *',
        'Баркод': 'Штрихкод *',
        'Штрихкод': 'Штрихкод *',
        'Высота упаковки': 'Габариты с упаковкой, см',  # Требует конвертации
        'Длина упаковки': 'Габариты с упаковкой, см',  # Требует конвертации  
        'Ширина упаковки': 'Габариты с упаковкой, см',  # Требует конвертации
        'Комплектация': 'Дополнительная информация',
        'Страна производства': 'Страна производства',
        'Материал изделия': 'Материал корыта',
        'Количество колес': 'Количество колес',
        'Диаметр колес': 'Диаметр колеса, см',  # Требует конвертации мм -> см
        'Грузоподъемность': 'Грузоподъемность, кг',
        'Объем': 'Объем, л',
        'Артикул WB': 'Артикул производителя'
    }
    
    # Словарь Яндекс -> WB
    yandex_to_wb = {
        'Ваш SKU *': 'Артикул продавца',
        'Название товара *': 'Наименование',
        'Бренд *': 'Бренд',
        'Описание товара *': 'Описание',
        'Ссылка на изображение *': 'Фото',
        'Ссылка на видео': 'Видео',
        'Вес с упаковкой, кг': 'Вес с упаковкой (кг)',
        'Название цвета от производителя': 'Цвет',
        'Цвет для фильтра': 'Цвет',
        'Цена *': 'Цена',
        'Штрихкод *': ['Баркод', 'Штрихкод'],
        'Габариты с упаковкой, см': ['Высота упаковки', 'Длина упаковки', 'Ширина упаковки'],  # Требует разделения
        'Дополнительная информация': 'Комплектация',
        'Страна производства': 'Страна производства',
        'Материал корыта': 'Материал изделия',
        'Количество колес': 'Количество колес',
        'Диаметр колеса, см': 'Диаметр колес',  # Требует конвертации см -> мм
        'Грузоподъемность, кг': 'Грузоподъемность',
        'Объем, л': 'Объем',
        'Артикул производителя': 'Артикул WB'
    }
    
    # Словарь Ozon -> Яндекс
    ozon_to_yandex = {
        'Артикул*': 'Ваш SKU *',
        'Название товара*': 'Название товара *',
        'Бренд*': 'Бренд *',
        'Аннотация': 'Описание товара *',
        'Ссылка на главное фото*': 'Ссылка на изображение *',
        'Ссылки на фото 360': 'Изображение для миниатюры',
        'Ссылки на видео': 'Ссылка на видео',
        'Вес в упаковке, г*': 'Вес с упаковкой, кг',  # Требует конвертации г -> кг
        'Цвет товара*': ['Название цвета от производителя', 'Цвет для фильтра'],
        'Цена, руб.*': 'Цена *',
        'Штрихкод (Серийный номер / EAN)': 'Штрихкод *',
        'Высота упаковки, мм*': 'Габариты с упаковкой, см',  # Требует конвертации мм -> см
        'Длина упаковки, мм*': 'Габариты с упаковкой, см',  # Требует конвертации мм -> см
        'Ширина упаковки, мм*': 'Габариты с упаковкой, см',  # Требует конвертации мм -> см
        'Комплектация': 'Дополнительная информация',
        'Страна-изготовитель': 'Страна производства',
        'Материал': 'Материал корыта',
        'Количество колес тачки': 'Количество колес',
        'Диаметр колеса, мм': 'Диаметр колеса, см',  # Требует конвертации мм -> см
        'Макс. нагрузка, кг': 'Грузоподъемность, кг',
        'Объем, л': 'Объем, л',
        'Хештеги': 'Теги'
    }
    
    # Словарь Ozon -> ЛеманПро
    ozon_to_lemanpro = {
        'Артикул*': 'Артикул товара',
        'Название товара*': 'Наименование товара мерчанта',
        'Бренд*': 'Бренд товара',
        'Название модели (для объединения в одну карточку)*': 'Модель товара',
        'Аннотация': 'Описание',
        'Вес в упаковке, г*': 'Вес в упаковке (кг)',  # Требует конвертации г -> кг
        'Цвет товара*': 'Цветовая палитра',
        'Штрихкод (Серийный номер / EAN)': 'Штрих-код',
        'Высота упаковки, мм*': 'Размеры в упаковке: высота (мм)',
        'Длина упаковки, мм*': 'Размеры в упаковке: длина (мм)',
        'Ширина упаковки, мм*': 'Размеры в упаковке: ширина (мм)',
        'Страна-изготовитель': 'Страна производства',
        'Материал': 'Основной материал',
        'ТН ВЭД коды ЕАЭС': 'Код ТН ВЭД',
        'Вес товара, г': 'Вес нетто (кг)',  # Требует конвертации г -> кг
        'Гарантийный срок': 'Гарантия (лет)',  # Может требовать преобразования формата
        'НДС, %*': 'НДС'
    }
    
    # Словарь ЛеманПро -> Ozon
    lemanpro_to_ozon = {
        'Артикул товара': 'Артикул*',
        'Наименование товара мерчанта': 'Название товара*',
        'Бренд товара': 'Бренд*',
        'Модель товара': 'Название модели (для объединения в одну карточку)*',
        'Описание': 'Аннотация',
        'Вес в упаковке (кг)': 'Вес в упаковке, г*',  # Требует конвертации кг -> г
        'Цветовая палитра': 'Цвет товара*',
        'Штрих-код': 'Штрихкод (Серийный номер / EAN)',
        'Размеры в упаковке: высота (мм)': 'Высота упаковки, мм*',
        'Размеры в упаковке: длина (мм)': 'Длина упаковки, мм*',
        'Размеры в упаковке: ширина (мм)': 'Ширина упаковки, мм*',
        'Страна производства': 'Страна-изготовитель',
        'Основной материал': 'Материал',
        'Код ТН ВЭД': 'ТН ВЭД коды ЕАЭС',
        'Вес нетто (кг)': 'Вес товара, г',  # Требует конвертации кг -> г
        'Гарантия (лет)': 'Гарантийный срок',  # Может требовать преобразования формата
        'НДС': 'НДС, %*'
    }
    
    # Поля, которые не должны переноситься
    excluded_columns = ['Артикул WB']
    
    mapping = {}
    used_target_columns = set()
    
    # Улучшенная нормализация имен колонок
    def normalize_column_name(col_name):
        if not isinstance(col_name, str):
            return str(col_name).lower()
        
        # Сохраняем исходное имя колонки для дальнейшего анализа
        original_name = col_name
        
        # Удаляем специальные маркеры из заголовков маркетплейсов: звездочки, восклицательные знаки и т.д.
        normalized = re.sub(r'[*!№\+]', '', col_name)
        
        # Удаляем спецсимволы и лишние пробелы
        normalized = re.sub(r'[^\w\s\-\.]', ' ', normalized)
        normalized = re.sub(r'\s+', ' ', normalized).strip().lower()
        
        # Обрабатываем многострочные заголовки (перенос строки заменяем на пробел)
        normalized = re.sub(r'\n', ' ', normalized)
        
        # Удаляем распространенные суффиксы
        common_suffixes = [
            " товара", " продукта", " позиции", " изделия", " шт", " г", " кг", " мл", " л",
            " см", " мм", " м", " руб", " rub", " ₽", " %", " руб."
        ]
        for suffix in common_suffixes:
            if normalized.endswith(suffix.lower()):
                normalized = normalized[:-len(suffix)]
        
        # Удаляем общие префиксы и пояснения в скобках
        normalized = re.sub(r'\([^)]*\)', '', normalized).strip()
        common_prefixes = [
            "код ", "номер ", "ид ", "тип ", "название ", "наименование ", "цена ", "стоимость ",
            "размер ", "вес ", "масса ", "ширина ", "высота ", "глубина ", "длина "
        ]
        for prefix in common_prefixes:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):]
        
        # Обрабатываем заголовки с единицами измерения через запятую
        normalized = re.sub(r',\s*(шт|г|кг|мл|л|см|мм|м|руб|rub|₽|%)$', '', normalized)
        
        # Преобразуем сокращения в полные формы
        abbreviations = {
            "артик": "артикул",
            "наим": "название",
            "наимен": "название",
            "описан": "описание",
            "кол-во": "количество",
            "кол во": "количество",
            "колво": "количество",
            "кол": "количество",
            "хар-ки": "характеристики",
            "хар ки": "характеристики",
            "харки": "характеристики",
            "хар-ка": "характеристика",
            "хар ка": "характеристика",
            "харка": "характеристика",
            "спец": "спецификация",
            "габар": "габариты",
            "разм": "размер",
            "фото": "изображение",
            "изобр": "изображение",
            "изобрa": "изображение",
            "картин": "изображение",
            "шир": "ширина",
            "дл": "длина",
            "выс": "высота",
            "глуб": "глубина"
        }
        for abbr, full in abbreviations.items():
            if normalized == abbr or normalized.startswith(abbr + " "):
                normalized = normalized.replace(abbr, full, 1)
                break
                
        # Стандартизация распространенных названий колонок для маркетплейсов
        marketplace_columns = {
            # Ключевые атрибуты
            "артикул": ["арт", "артик", "код товара", "номер артикула", "skuмагазина", "sku", "id товара", 
                      "код позиции", "wb sku", "артикул продавца"],
            "название": ["наименование", "имя", "наимен", "назв", "имя товара", "заголовок", "title", 
                       "наименование товара", "название товара", "наименование позиции", "полное название"],
            "цена": ["стоимость", "розн цена", "цена продажи", "price", "розничная цена", "прайс", 
                    "цена товара", "цена со скидкой", "розничная", "цена розничная", "руб"],
            "описание": ["описание товара", "полное описание", "detail", "детальное описание", "description", 
                        "расширенное описание", "контент", "content", "информация о товаре", "товар описание"],
            "категория": ["раздел", "группа", "группа товаров", "category", "тип товара", "тип изделия", 
                        "категория товара", "родительская категория", "товарная категория", "предметная группа"],
            "бренд": ["брэнд", "марка", "производитель", "brand", "изготовитель", "торговая марка", "тм", 
                     "товарный знак", "компания производитель", "марка производитель"],
            
            # Габаритные характеристики
            "вес": ["масса", "вес товара", "вес в упаковке", "вес без упаковки", "вес брутто", "вес нетто", 
                   "weight", "масса товара", "масса в упаковке", "объемный вес"],
            "ширина": ["width", "ширина товара", "ширина упаковки", "ширина изделия", "ширина габарит",
                      "ширина в упаковке", "ширина без упаковки", "габариты ширина"],
            "высота": ["height", "высота товара", "высота упаковки", "высота изделия", "высота габарит",
                      "высота в упаковке", "высота без упаковки", "габариты высота"],
            "длина": ["length", "глубина", "длина товара", "длина упаковки", "длина изделия", "длина габарит",
                     "длина в упаковке", "длина без упаковки", "габариты длина", "глубина габарит"],
            
            # Логистика и наличие
            "количество": ["кол-во", "остаток", "остатки", "наличие", "колво", "qty", "quantity", 
                          "количество штук", "доступное количество", "количество в наличии"],
            "баркод": ["штрихкод", "шк", "баркод товара", "ean", "ean13", "gtin", "upc", "код товара", 
                      "штрих код", "barcode", "штрихкод товара"],
            
            # Дополнительные атрибуты
            "материал": ["состав", "материал изготовления", "материал товара", "материал изделия", 
                        "основной материал", "material", "ткань", "основа", "сырье"],
            "цвет": ["color", "расцветка", "цвет товара", "цвет изделия", "основной цвет", 
                    "цветовой тон", "оттенок", "цвет и оттенок", "цветовое решение"],
            "размер": ["габариты", "size", "размерный ряд", "размер товара", "размер изделия", 
                      "линейные размеры", "типоразмер", "размерность", "габаритные размеры"],
                    
            # Фотографии
            "фото": ["изображение", "картинка", "главное фото", "основное фото", "фотография", "photo", "image", 
                    "ссылка на фото", "фото товара", "фотки", "снимки", "изображения", "ссылки на фото", 
                    "picture", "ссылка на изображение", "ссылки на изображения", "ссылка на главное фото", 
                    "ссылка на основное фото", "ссылки на фотографии", "ссылки на картинки", "фото 360"],
            "главное фото": ["ссылка на главное фото", "основное фото", "главное изображение", "main photo", 
                           "main image", "первое фото", "фото основное", "фото в карточке", "основная картинка", 
                           "основное изображение", "главная фотография", "главное фото товара"],
            "дополнительные фото": ["ссылки на дополнительные фото", "дополнительные изображения", "additional photos", 
                                 "дополнительные картинки", "доп. фото", "доп фото", "галерея", "gallery", 
                                 "фото галерея", "дополнительные фотографии", "фото товара"]
        }
        
        # Проверяем, соответствует ли нормализованное имя одному из стандартных имён
        for standard, aliases in marketplace_columns.items():
            if normalized == standard or normalized in aliases:
                return standard
                
        # Для оригинальных имен колонок с маркерами обязательности (звездочка и др.)
        # добавляем повышенный приоритет для ключевых атрибутов
        if '*' in original_name or '!' in original_name:
            for key_field, aliases in {'артикул': marketplace_columns['артикул'], 
                                      'название': marketplace_columns['название'], 
                                      'цена': marketplace_columns['цена']}.items():
                for alias in aliases:
                    if alias in normalized or normalized in alias:
                        return key_field
                        
        # Если колонка содержит цифры (например, Артикул1, Артикул2), очищаем от них
        normalized = re.sub(r'\d+$', '', normalized).strip()
                
        return normalized
    
    # Функция для определения типа маркетплейса только по столбцам
    def detect_marketplace_from_columns(columns):
        if not columns:
            return 'unknown'
            
        # Характерные колонки для Wildberries
        wb_indicators = ['Артикул продавца', 'Артикул WB', 'Номер номенклатуры', 'Наименование']
        
        # Характерные колонки для Ozon с маркером обязательности (*)
        ozon_indicators = ['Артикул*', 'Название товара*', 'Бренд*', 'Ссылка на главное фото*', 
                          'Вес в упаковке, г*', 'Высота упаковки, мм*', 'Длина упаковки, мм*', 'Ширина упаковки, мм*']
        
        wb_matches = sum(1 for col in columns if col in wb_indicators)
        ozon_matches = sum(1 for col in columns if col in ozon_indicators)
        
        if wb_matches > ozon_matches:
            return 'wildberries'
        elif ozon_matches > wb_matches:
            return 'ozon'
        else:
            # Если не удалось однозначно определить, возвращаем 'unknown'
            return 'unknown'
    
    # Определяем тип маркетплейса для исходных и целевых колонок
    source_marketplace = detect_marketplace_from_columns(source_columns)
    target_marketplace = detect_marketplace_from_columns(target_columns)
    
    # Применяем готовые словари соответствия в зависимости от направления
    if source_marketplace == 'wildberries' and target_marketplace == 'ozon':
        # WB -> Ozon
        for source_col in source_columns:
            if source_col in wb_to_ozon and wb_to_ozon[source_col] in target_columns:
                mapping[source_col] = wb_to_ozon[source_col]
                used_target_columns.add(wb_to_ozon[source_col])
    elif source_marketplace == 'ozon' and target_marketplace == 'wildberries':
        # Ozon -> WB
        for source_col in source_columns:
            if source_col in ozon_to_wb and ozon_to_wb[source_col] in target_columns:
                mapping[source_col] = ozon_to_wb[source_col]
                used_target_columns.add(ozon_to_wb[source_col])
    
    # Проверка схожести названий колонок для оставшихся колонок
    for source_col in source_columns:
        if source_col not in mapping:
            source_normalized = normalize_column_name(source_col)
            
            best_match = None
            best_score = 0
            
            for target_col in target_columns:
                if target_col not in used_target_columns:
                    target_normalized = normalize_column_name(target_col)
                    
                    # Проверяем нечеткое совпадение
                    score = fuzz.token_sort_ratio(source_normalized, target_normalized)
                    
                    # Повышаем вес для точных соответствий
                    if source_normalized == target_normalized:
                        score += 30
                    
                    # Повышаем вес для частичных соответствий
                    elif source_normalized in target_normalized or target_normalized in source_normalized:
                        score += 15
                    
                    # Обрабатываем специфические отношения между единицами измерения
                    # Вес: кг <-> г
                    if ('кг' in source_col.lower() and 'г' in target_col.lower() and 'кг' not in target_col.lower()) or \
                       ('кг' in target_col.lower() and 'г' in source_col.lower() and 'кг' not in source_col.lower()):
                        if any(w in source_normalized for w in ['вес', 'масса']) and \
                           any(w in target_normalized for w in ['вес', 'масса']):
                            score += 20
                    
                    # Размеры: мм <-> см
                    if ('мм' in source_col.lower() and 'см' in target_col.lower()) or \
                       ('мм' in target_col.lower() and 'см' in source_col.lower()):
                        if any(w in source_normalized for w in ['длина', 'ширина', 'высота', 'глубина', 'габарит', 'размер']) and \
                           any(w in target_normalized for w in ['длина', 'ширина', 'высота', 'глубина', 'габарит', 'размер']):
                            score += 20
                    
                    if score > best_score and score >= threshold:
                        best_score = score
                        best_match = target_col
            
            if best_match:
                mapping[source_col] = best_match
                used_target_columns.add(best_match)
    
    return mapping

def transfer_data_between_tables(source_df, target_workbook, target_sheet_name, column_mapping, target_header_row=1, source_filename=None):
    """
    Переносит данные из исходного DataFrame в целевую таблицу, сохраняя форматирование
    
    Args:
        source_df: DataFrame с исходными данными
        target_workbook: Объект целевой рабочей книги openpyxl
        target_sheet_name: Имя целевого листа
        column_mapping: Словарь соответствия колонок {source_column: target_column}
        target_header_row: Номер строки с заголовками в целевой таблице (по умолчанию 1)
        source_filename: Имя исходного файла (для заполнения поля "Категория продавца")
        
    Returns:
        Объект рабочей книги openpyxl с обновленными данными
    """
    target_sheet = target_workbook[target_sheet_name]
    header_row = target_header_row
    
    # Определение колонок в целевой таблице
    target_column_indices = {}
    for col_idx in range(1, target_sheet.max_column + 1):
        cell = target_sheet.cell(row=header_row, column=col_idx)
        if cell.value:
            target_column_indices[str(cell.value)] = col_idx
    
    # Проверка наличия подзаголовков в целевой таблице (строка сразу после заголовков)
    has_subheaders = False
    subheader_info = {}
    
    # Проверяем наличие данных в строке после заголовков
    subheader_row = target_sheet[header_row + 1]
    subheader_values = [cell.value for cell in subheader_row if cell.value]
    if subheader_values:
        has_subheaders = True
        # Сохраняем информацию о подзаголовках и их форматировании
        for col_name, col_idx in target_column_indices.items():
            subheader_cell = target_sheet.cell(row=header_row + 1, column=col_idx)
            subheader_info[col_name] = {
                'value': subheader_cell.value,
                'font': subheader_cell.font.copy() if subheader_cell.font else None,
                'fill': subheader_cell.fill.copy() if subheader_cell.fill else None,
                'border': subheader_cell.border.copy() if subheader_cell.border else None,
                'alignment': subheader_cell.alignment.copy() if subheader_cell.alignment else None,
                'number_format': subheader_cell.number_format,
                'protection': subheader_cell.protection.copy() if subheader_cell.protection else None
            }
    
    # Определяем, с какой строки начинаются данные в исходной таблице
    has_source_subheaders = False
    if len(source_df) > 0:
        first_row = source_df.iloc[0]
        string_descriptors = 0
        numeric_values = 0
        
        for col in source_df.columns:
            val = first_row[col]
            if isinstance(val, str) and not any(c.isdigit() for c in val):
                string_descriptors += 1
            elif isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
                
        # Если в первой строке больше нечисловых описательных значений, это может быть подзаголовок
        has_source_subheaders = string_descriptors > numeric_values
    
    # Определяем начальный индекс для данных в исходной таблице
    data_start_idx = 1 if has_source_subheaders else 0
    
    # Определяем начальный индекс для данных в целевой таблице
    target_data_start_row = header_row + 2 if has_subheaders else header_row + 1
    
    # Сохраняем образец форматирования для каждой колонки целевой таблицы
    # Используем первую строку с данными, если она есть
    style_info = {}
    if target_sheet.max_row >= target_data_start_row:
        for col_name, col_idx in target_column_indices.items():
            style_cell = target_sheet.cell(row=target_data_start_row, column=col_idx)
            style_info[col_name] = {
                'font': style_cell.font.copy() if style_cell.font else None,
                'fill': style_cell.fill.copy() if style_cell.fill else None,
                'border': style_cell.border.copy() if style_cell.border else None,
                'alignment': style_cell.alignment.copy() if style_cell.alignment else None,
                'number_format': style_cell.number_format,
                'protection': style_cell.protection.copy() if style_cell.protection else None
            }
    
    # Очищаем данные в целевой таблице (оставляем заголовки и подзаголовки)
    for row_idx in range(target_data_start_row, target_sheet.max_row + 1):
        for col_idx in range(1, target_sheet.max_column + 1):
            target_sheet.cell(row=row_idx, column=col_idx).value = None
    
    # Копируем данные из исходной таблицы
    if len(source_df) > data_start_idx:
        data_to_copy = source_df.iloc[data_start_idx:]
        
        for idx, source_row in data_to_copy.iterrows():
            target_row_idx = target_data_start_row + (idx - data_start_idx)
            
            for source_col, target_col_value in column_mapping.items():
                # Проверяем, не входит ли колонка в список исключений
                if source_col in excluded_columns:
                    continue
                    
                # Обрабатываем случай, когда target_col - это список
                if isinstance(target_col_value, list):
                    target_cols = target_col_value
                else:
                    target_cols = [target_col_value]
                    
                # Проходим по всем целевым колонкам
                for target_col in target_cols:
                    # Проверяем, существуют ли исходная и целевая колонки, и не в списке ли исключений целевая колонка
                    if source_col in source_df.columns and target_col in target_column_indices and target_col not in excluded_columns:
                        # Получаем значение из исходной таблицы
                        value = source_row.get(source_col)
                        target_col_idx = target_column_indices[target_col]
                        
                        # Предобработка значения и безопасное преобразование
                        if value is not None and not pd.isna(value):
                            # Для строк и чисел - разные преобразования
                            try:
                                # Преобразование единиц измерения
                                source_col_lower = source_col.lower()
                                target_col_lower = target_col.lower()
                                
                                # Для чисел, представленных как строки, пробуем безопасно конвертировать
                                if isinstance(value, str) and (value.replace('.', '', 1).replace(',', '', 1).isdigit() or 
                                                              value.lstrip('-').replace('.', '', 1).replace(',', '', 1).isdigit()):
                                    # Заменяем запятую на точку для правильного преобразования
                                    numeric_value = float(value.replace(',', '.'))
                                    
                                    # Вес: кг -> г
                                    if (('вес' in source_col_lower or 'масса' in source_col_lower) and 
                                       'кг' in source_col_lower and 'г' in target_col_lower and 'кг' not in target_col_lower):
                                        value = numeric_value * 1000  # Из кг в г
                                    
                                    # Вес: г -> кг
                                    elif (('вес' in source_col_lower or 'масса' in source_col_lower) and 
                                          'г' in source_col_lower and 'кг' not in source_col_lower and 'кг' in target_col_lower):
                                        value = numeric_value / 1000  # Из г в кг
                                    
                                    # Размеры: мм -> см
                                    elif ((any(dim in source_col_lower for dim in ['длина', 'ширина', 'высота', 'глубина', 'диаметр']) or 
                                         'упаковк' in source_col_lower) and 
                                          'мм' in source_col_lower and 'см' in target_col_lower):
                                        value = numeric_value / 10  # Из мм в см
                                    
                                    # Размеры: см -> мм
                                    elif ((any(dim in source_col_lower for dim in ['длина', 'ширина', 'высота', 'глубина', 'диаметр']) or 
                                         'упаковк' in source_col_lower) and 
                                          'см' in source_col_lower and 'мм' in target_col_lower):
                                        value = numeric_value * 10  # Из см в мм
                                    else:
                                        # Если не нужно преобразовывать единицы, но значение числовое
                                        # оставляем его в виде строки для избежания проблем с конвертацией
                                        # например для полей типа "Артикул", "SKU" и т.д.
                                        if any(id_field in target_col_lower for id_field in ['sku', 'артикул', 'guid', 'штрихкод', 'баркод']):
                                            value = str(value)  # Оставляем как строку
                                        else:
                                            value = numeric_value  # Используем числовое значение
                                            
                                # Для числовых значений
                                elif isinstance(value, (int, float)):
                                    # Вес: кг -> г
                                    if (('вес' in source_col_lower or 'масса' in source_col_lower) and 
                                       'кг' in source_col_lower and 'г' in target_col_lower and 'кг' not in target_col_lower):
                                        value = value * 1000  # Из кг в г
                                    
                                    # Вес: г -> кг
                                    elif (('вес' in source_col_lower or 'масса' in source_col_lower) and 
                                          'г' in source_col_lower and 'кг' not in source_col_lower and 'кг' in target_col_lower):
                                        value = value / 1000  # Из г в кг
                                    
                                    # Размеры: мм -> см
                                    elif ((any(dim in source_col_lower for dim in ['длина', 'ширина', 'высота', 'глубина', 'диаметр']) or 
                                         'упаковк' in source_col_lower) and 
                                          'мм' in source_col_lower and 'см' in target_col_lower):
                                        value = value / 10  # Из мм в см
                                    
                                    # Размеры: см -> мм
                                    elif ((any(dim in source_col_lower for dim in ['длина', 'ширина', 'высота', 'глубина', 'диаметр']) or 
                                         'упаковк' in source_col_lower) and 
                                          'см' in source_col_lower and 'мм' in target_col_lower):
                                        value = value * 10  # Из см в мм
                                        
                                # Для идентификаторов (SKU, Артикул и т.д.) - всегда преобразуем в строку
                                if any(id_field in target_col_lower for id_field in ['sku', 'артикул', 'код товара', 'guid', 'штрихкод', 'баркод']):
                                    value = str(value)
                            except (ValueError, TypeError) as e:
                                # В случае ошибки преобразования оставляем исходное значение
                                pass
                                    
                        # Специальная обработка для поля "Категория продавца" - использование имени файла
                        # Заполняем поле "Категория продавца" всегда, когда оно есть в шаблоне
                        if target_col == "Категория продавца" and source_filename:
                            # Извлекаем название файла без расширения
                            filename_without_ext = os.path.splitext(os.path.basename(source_filename))[0]
                            value = filename_without_ext
                            
                        # Специальная обработка для фотографий - здесь i это индекс строки
                        i = idx - data_start_idx  # Индекс строки в исходном DataFrame
                        
                        # WB -> Ozon: поле "Фото" переносится в "Ссылка на главное фото*" (первая ссылка) и "Ссылки на дополнительные фото"
                        if source_col == "Фото" and (target_col == "Ссылка на главное фото*" or target_col == "Ссылки на дополнительные фото"):
                            if value and isinstance(value, str):
                                # Разделяем строку с фотографиями по точке с запятой (характерно для WB)
                                photo_links = []
                                if ';' in value:
                                    photo_links = [link.strip() for link in value.split(';') if link.strip()]
                                # Если нет разделителя точка с запятой, используем другие разделители
                                else:
                                    photo_links = [link.strip() for link in re.split(r'[\n\r,;]+', value.strip()) if link.strip()]
                                
                                # Очищаем ссылки и оставляем только URL
                                cleaned_links = []
                                for link in photo_links:
                                    if link.startswith('http'):
                                        cleaned_links.append(link)
                                    else:
                                        # Ищем URL в строке
                                        urls = re.findall(r'https?://[^\s,;]+', link)
                                        cleaned_links.extend(urls)
                                
                                # Если нашли URL
                                if cleaned_links:
                                    if target_col == "Ссылка на главное фото*":
                                        # Берем только первую ссылку для главного фото
                                        value = cleaned_links[0]
                                    elif target_col == "Ссылки на дополнительные фото":
                                        # Для дополнительных фото берем все кроме первой, добавляем каждую с новой строки без разделителя ";"
                                        if len(cleaned_links) > 1:
                                            # Используем перенос строки (без других разделителей)
                                            value = '\n'.join(cleaned_links[1:])
                                        else:
                                            value = ""  # Если есть только одно фото, то поле дополнительных фото оставляем пустым
                                        
                        # Ozon -> WB: объединяем "Ссылка на главное фото*" и "Ссылки на дополнительные фото" в поле "Фото"
                        if (source_col == "Ссылка на главное фото*" or source_col == "Ссылки на дополнительные фото") and target_col == "Фото":
                            # Если мы обрабатываем главное фото
                            if source_col == "Ссылка на главное фото*":
                                # Преобразуем значение в строку, чтобы избежать ошибок конкатенации
                                main_photo = str(value) if value else ""
                                
                                # Ищем дополнительные фото в этой же строке
                                additional_photos_col = "Ссылки на дополнительные фото"
                                if additional_photos_col in source_df.columns:
                                    additional_photos = source_row.get(additional_photos_col, "")
                                    if additional_photos and isinstance(additional_photos, str):
                                        # Разбиваем дополнительные фото, которые могут быть с переносом строки
                                        add_photos_split = re.split(r'[\n\r,;]+', additional_photos.strip())
                                        add_photos_clean = [p.strip() for p in add_photos_split if p and p.strip().startswith('http')]
                                        
                                        # Объединяем фото с правильным разделителем - точка с запятой для WB
                                        if main_photo:
                                            all_photos = [main_photo] + add_photos_clean
                                            value = ';'.join(all_photos)
                                        else:
                                            value = ';'.join(add_photos_clean) if add_photos_clean else ""
                            
                            # Если мы обрабатываем дополнительные фото, то пропускаем обработку,
                            # так как они уже обработаны вместе с главным фото
                            elif source_col == "Ссылки на дополнительные фото":
                                # Проверяем, обрабатывали ли мы уже эту пару колонок через главное фото
                                main_photo_col = "Ссылка на главное фото*"
                                if main_photo_col in source_df.columns and main_photo_col in column_mapping:
                                    continue  # Пропускаем, так как эти данные уже должны быть обработаны через главное фото
                        
                        # Записываем значение в целевую таблицу
                        cell = target_sheet.cell(row=target_row_idx, column=target_col_idx)
                        cell.value = value
                    
                    # Применяем сохраненное форматирование из образца данных (не из подсказок)
                    if target_col in style_info:
                        cell_style = style_info[target_col]
                        if cell_style['font']: cell.font = cell_style['font']
                        if cell_style['fill']: cell.fill = cell_style['fill']
                        if cell_style['border']: cell.border = cell_style['border']
                        if cell_style['alignment']: cell.alignment = cell_style['alignment']
                        if cell_style['number_format']: cell.number_format = cell_style['number_format']
                        if cell_style['protection']: cell.protection = cell_style['protection']
    
    # Восстанавливаем подзаголовки в целевой таблице, если они были
    if has_subheaders:
        for col_name, col_idx in target_column_indices.items():
            if col_name in subheader_info:
                subheader_cell = target_sheet.cell(row=header_row + 1, column=col_idx)
                subheader_cell.value = subheader_info[col_name]['value']
                
                # Восстанавливаем форматирование подзаголовка
                subheader_style = subheader_info[col_name]
                if subheader_style['font']: subheader_cell.font = subheader_style['font']
                if subheader_style['fill']: subheader_cell.fill = subheader_style['fill']
                if subheader_style['border']: subheader_cell.border = subheader_style['border']
                if subheader_style['alignment']: subheader_cell.alignment = subheader_style['alignment']
                if subheader_style['number_format']: subheader_cell.number_format = subheader_style['number_format']
                if subheader_style['protection']: subheader_cell.protection = subheader_style['protection']
    
    return target_workbook

def preview_data(source_df, target_df, column_mapping, source_filename=None):
    """
    Создает предварительный просмотр того, как данные будут выглядеть после переноса
    
    Args:
        source_df: DataFrame с исходными данными
        target_df: DataFrame целевой таблицы
        column_mapping: Словарь соответствия колонок {source_column: target_column}
        source_filename: Имя исходного файла (для заполнения поля "Категория продавца")
        
    Returns:
        DataFrame: DataFrame с предварительным просмотром
    """
    # Создаем копию целевого DataFrame
    preview_df = target_df.copy()
    
    # Убедимся, что колонки Артикул WB нет в DataFrame или она представлена как строка
    # Это предотвратит ошибку конвертации Arrow
    if 'Артикул WB' in preview_df.columns:
        preview_df['Артикул WB'] = preview_df['Артикул WB'].astype(str)
    
    # Проверяем наличие подзаголовков в исходной таблице
    has_source_subheaders = False
    if len(source_df) > 0:
        first_row = source_df.iloc[0]
        numeric_values = 0
        string_descriptors = 0
        
        for col in source_df.columns:
            val = first_row[col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
            elif isinstance(val, str) and not any(c.isdigit() for c in val):
                string_descriptors += 1
        
        # Если в первой строке больше нечисловых описательных значений, это может быть подзаголовок
        has_source_subheaders = string_descriptors > numeric_values
    
    # Фильтруем данные, пропуская подзаголовки если они есть
    data_start_idx = 1 if has_source_subheaders else 0
    filtered_source_df = source_df.iloc[data_start_idx:].copy() if len(source_df) > data_start_idx else source_df.copy()
    
    # Используем глобальную переменную excluded_columns, которая определена в начале файла
    
    # Создаем соответствие между колонками источника и целевой таблицы
    col_pairs = []
    for src, tgt_value in column_mapping.items():
        # Пропускаем исключенные колонки
        if src in excluded_columns:
            continue
            
        # Обрабатываем случай, когда target_col - это список
        if isinstance(tgt_value, list):
            target_cols = tgt_value
        else:
            target_cols = [tgt_value]
            
        # Добавляем пары для каждой целевой колонки
        for tgt in target_cols:
            if tgt in preview_df.columns and tgt not in excluded_columns:
                col_pairs.append((src, tgt))
    
    # Проверяем наличие подзаголовков в целевой таблице
    has_target_subheaders = False
    if len(preview_df) > 0:
        first_row = preview_df.iloc[0]
        numeric_values = 0
        string_descriptors = 0
        
        for col in preview_df.columns:
            val = first_row[col]
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values += 1
            elif isinstance(val, str) and not any(c.isdigit() for c in val) and not pd.isna(val):
                string_descriptors += 1
        
        has_target_subheaders = string_descriptors > numeric_values
    
    # Создаем новый DataFrame для предпросмотра
    result_columns = preview_df.columns
    result_df = pd.DataFrame(columns=result_columns)
    
    # Сохраняем подзаголовки из целевой таблицы, если они есть
    if has_target_subheaders and len(preview_df) > 0:
        result_df = pd.concat([result_df, preview_df.iloc[[0]]], ignore_index=True)
        target_data_start = 1
    else:
        target_data_start = 0
    
    # Копируем данные из исходной таблицы в соответствующие колонки целевой
    for idx, source_row in filtered_source_df.iterrows():
        new_row = {col: None for col in result_columns}
        
        # Специальная обработка для поля "Категория продавца" - используем имя файла
        # Заполняем поле "Категория продавца" всегда, когда оно есть в результирующей таблице
        if "Категория продавца" in result_columns and source_filename:
            filename_without_ext = os.path.splitext(os.path.basename(source_filename))[0]
            new_row["Категория продавца"] = filename_without_ext
        
        for src, tgt in col_pairs:
            if src in filtered_source_df.columns:
                value = source_row.get(src)
                
                # Обработка значения для совместимости
                if value is not None:
                    # Если это число, оставляем как есть
                    if isinstance(value, (int, float)):
                        pass
                    # Для других типов - преобразуем в строку
                    elif pd.isna(value):
                        value = None
                    elif not isinstance(value, str):
                        value = str(value)
                
                # Специальная обработка для фотографий
                # WB -> Ozon: поле "Фото" переносится в "Ссылка на главное фото*" и "Ссылки на дополнительные фото"
                if src == "Фото" and (tgt == "Ссылка на главное фото*" or tgt == "Ссылки на дополнительные фото"):
                    if value and isinstance(value, str):
                        # Разделяем строку с фотографиями по точке с запятой (характерно для WB)
                        photo_links = []
                        if ';' in value:
                            photo_links = [link.strip() for link in value.split(';') if link.strip()]
                        # Если нет разделителя точка с запятой, используем другие разделители
                        else:
                            photo_links = [link.strip() for link in re.split(r'[\n\r,;]+', value.strip()) if link.strip()]
                        
                        # Очищаем ссылки и оставляем только URL
                        cleaned_links = []
                        for link in photo_links:
                            if link.startswith('http'):
                                cleaned_links.append(link)
                            else:
                                # Ищем URL в строке
                                urls = re.findall(r'https?://[^\s,;]+', link)
                                cleaned_links.extend(urls)
                        
                        # Если нашли URL
                        if cleaned_links:
                            if tgt == "Ссылка на главное фото*":
                                # Берем только первую ссылку для главного фото
                                value = cleaned_links[0]
                            elif tgt == "Ссылки на дополнительные фото":
                                # Для дополнительных фото берем все кроме первой, добавляем каждую с новой строки без разделителя ";"
                                if len(cleaned_links) > 1:
                                    # Используем перенос строки (без других разделителей)
                                    value = '\n'.join(cleaned_links[1:])
                                else:
                                    value = ""  # Если есть только одно фото, то поле дополнительных фото оставляем пустым
                
                # Ozon -> WB: объединяем "Ссылка на главное фото*" и "Ссылки на дополнительные фото" в поле "Фото"
                if (src == "Ссылка на главное фото*" or src == "Ссылки на дополнительные фото") and tgt == "Фото":
                    if src == "Ссылка на главное фото*":
                        # Преобразуем значение в строку, чтобы избежать ошибок конкатенации
                        main_photo = str(value) if value else ""
                        
                        # Ищем дополнительные фото в этой же строке
                        additional_photos_col = "Ссылки на дополнительные фото"
                        if additional_photos_col in filtered_source_df.columns:
                            additional_photos = source_row.get(additional_photos_col, "")
                            if additional_photos and isinstance(additional_photos, str):
                                # Разбиваем дополнительные фото, которые могут быть с переносом строки
                                add_photos_split = re.split(r'[\n\r,;]+', additional_photos.strip())
                                add_photos_clean = [p.strip() for p in add_photos_split if p and p.strip().startswith('http')]
                                
                                # Объединяем фото с правильным разделителем - точка с запятой для WB
                                if main_photo:
                                    all_photos = [main_photo] + add_photos_clean
                                    value = ';'.join(all_photos)
                                else:
                                    value = ';'.join(add_photos_clean) if add_photos_clean else ""
                    
                    # Если мы обрабатываем дополнительные фото, пропускаем - они уже обработаны вместе с главным фото
                    elif src == "Ссылки на дополнительные фото":
                        main_photo_col = "Ссылка на главное фото*"
                        if main_photo_col in filtered_source_df.columns:
                            # Проверяем, есть ли соответствие для главного фото
                            for s, t in col_pairs:
                                if s == main_photo_col and t == tgt:
                                    # Пропускаем, так как эти данные уже должны быть обработаны через главное фото
                                    value = None
                                    break
                
                new_row[tgt] = value
        
        # Добавляем строку с данными к результату
        result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)
    
    # Преобразуем все колонки в строковый тип для предотвращения проблем при отображении
    for col in result_df.columns:
        # Сначала конвертируем в строки те значения, что могут вызывать проблемы
        # Используем метод, который будет работать с разными типами данных
        result_df[col] = result_df[col].apply(lambda x: str(x) if x is not None and not pd.isna(x) else None)
    
    return result_df
