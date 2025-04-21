import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import io
import tempfile
import os
import re
import base64
from fuzzywuzzy import fuzz

# Импортируем новый модуль распознавания маркетплейсов
import marketplace_detection

# Функция для конвертации изображения в base64
def get_image_base64(image_path):
    """Преобразует изображение в строку base64 для отображения через HTML"""
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode('utf-8')
from utils import (
    load_excel_file, 
    save_excel_file, 
    map_columns_automatically, 
    transfer_data_between_tables,
    preview_data,
    find_header_row,
    detect_marketplace_template,
    find_best_marketplace_sheet
)

# Настройка страницы
st.set_page_config(
    page_title="Маппинг таблиц маркетплейсов",
    page_icon="📊",
    layout="wide"
)

# Инициализация состояний сессии
if 'source_file' not in st.session_state:
    st.session_state.source_file = None
if 'target_file' not in st.session_state:
    st.session_state.target_file = None
if 'source_data' not in st.session_state:
    st.session_state.source_data = None
if 'target_data' not in st.session_state:
    st.session_state.target_data = None
if 'source_columns' not in st.session_state:
    st.session_state.source_columns = None
if 'target_columns' not in st.session_state:
    st.session_state.target_columns = None
if 'column_mapping' not in st.session_state:
    st.session_state.column_mapping = {}
if 'mapping_complete' not in st.session_state:
    st.session_state.mapping_complete = False
if 'transfer_complete' not in st.session_state:
    st.session_state.transfer_complete = False
if 'auto_mapped' not in st.session_state:
    st.session_state.auto_mapped = False
if 'source_sheet_name' not in st.session_state:
    st.session_state.source_sheet_name = None
if 'target_sheet_name' not in st.session_state:
    st.session_state.target_sheet_name = None
if 'source_sheets' not in st.session_state:
    st.session_state.source_sheets = []
if 'target_sheets' not in st.session_state:
    st.session_state.target_sheets = []
if 'source_workbook' not in st.session_state:
    st.session_state.source_workbook = None
if 'target_workbook' not in st.session_state:
    st.session_state.target_workbook = None
if 'preview_result' not in st.session_state:
    st.session_state.preview_result = None
if 'source_header_row' not in st.session_state:
    st.session_state.source_header_row = 1
if 'target_header_row' not in st.session_state:
    st.session_state.target_header_row = 1

# Заголовок и описание
st.title("🔄 Маппинг таблиц маркетплейсов")
st.caption("Инструмент для переноса данных между шаблонами таблиц маркетплейсов")


st.divider()

# Основное содержимое: два столбца
col1, col2 = st.columns(2)

with col1:
    st.subheader("📤 Исходная таблица (Откуда)")
    
    # Добавляем логотипы маркетплейсов в области загрузки файлов в один ряд маленькими иконками
    st.write(
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/xlsx.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/wildberries.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/ozon.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/Яндекс маркет.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/Лемана про.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/все инструменты.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/сбермегамаркет.png")}" width="18" style="margin-right:4px"></span>',
        unsafe_allow_html=True
    )
    
    source_file = st.file_uploader("Загрузите исходную таблицу (xlsx)", type=['xlsx'], key="source_uploader")
    
    if source_file is not None and source_file != st.session_state.source_file:
        st.session_state.source_file = source_file
        try:
            source_workbook, source_sheets = load_excel_file(source_file)
            st.session_state.source_workbook = source_workbook
            st.session_state.source_sheets = source_sheets
            
            if len(source_sheets) > 0:
                # По умолчанию используем первый лист
                selected_sheet = source_sheets[0]
                header_row = 1
                
                # Маркетплейсы и характерные для них листы и строки заголовков
                marketplace_config = {
                    'ozon': {
                        'sheet_patterns': ['шаблон', 'template', 'озон', 'ozon'],
                        'header_row': 2
                    },
                    'wildberries': {
                        'sheet_patterns': ['товары', 'вб', 'wb', 'wildberries'],
                        'header_row': 3
                    },
                    'lemanpro': {
                        'sheet_patterns': ['леман', 'атем', 'leman', 'atem'],
                        'header_row': 4
                    },
                    'yandex': {
                        'sheet_patterns': ['данные о товарах', 'яндекс', 'маркет', 'яндекс маркет', 'yandex'],
                        'header_row': 2
                    }
                }
                
                # Шаг 1: Ищем подходящий лист по имени
                sheet_found = False
                
                # Специальная проверка для ЛеманПро по имени файла
                source_filename = getattr(source_file, 'name', '')
                if source_filename:
                    # Для ЛеманПро проверяем соответствие имени файла и названия листа
                    file_base_name = os.path.splitext(os.path.basename(source_filename))[0]
                    
                    for sheet_name in source_sheets:
                        # Ищем лист, в названии которого есть "шаблон" и имя файла
                        # или характерные для ЛеманПро имена
                        if ("шаблон" in sheet_name.lower() and file_base_name.lower() in sheet_name.lower()) or \
                           any(pattern in sheet_name.lower() for pattern in ["атем", "atem", "леман"]):
                            selected_sheet = sheet_name
                            header_row = 4  # Для ЛеманПро характерна 4-я строка заголовков
                            marketplace_type = 'lemanpro'
                            sheet_found = True
                            break
                
                # Если не нашли ЛеманПро, ищем другие маркетплейсы
                if not sheet_found:
                    for sheet_name in source_sheets:
                        sheet_lower = sheet_name.lower()
                        for marketplace, config in marketplace_config.items():
                            if any(pattern in sheet_lower for pattern in config['sheet_patterns']):
                                selected_sheet = sheet_name
                                header_row = config['header_row']
                                marketplace_type = marketplace
                                sheet_found = True
                                break
                        if sheet_found:
                            break
                
                # Устанавливаем лист и строку заголовков
                st.session_state.source_sheet_name = selected_sheet
                st.session_state.source_header_row = header_row
                
                # Шаг 2: Получаем заголовки из выбранного листа
                sheet = source_workbook[selected_sheet]
                if sheet.max_row >= header_row:
                    # Собираем заголовки
                    headers = []
                    for cell in sheet[header_row]:
                        if cell.value is not None and str(cell.value).strip() != "":
                            headers.append(str(cell.value))
                    
                    # Шаг 3: Определяем маркетплейс по заголовкам с учетом строки и первых 5 колонок
                    if headers:
                        # Используем новый модуль распознавания для более точного определения
                        normalized_columns = [str(h).lower().strip() for h in headers]
                        marketplace_type, confidence, details = marketplace_detection.detect_marketplace_by_row_headers(
                            normalized_columns, 
                            st.session_state.source_header_row
                        )
                        
                        # Проверяем, есть ли в названии листа "Данные" для "Все инструменты"
                        sheet_name_lower = selected_sheet.lower()
                        if 'данные' in sheet_name_lower and st.session_state.source_header_row == 2:
                            # Проверяем на шаблон Все инструменты
                            has_guid = any('guid*' in col.lower() for col in normalized_columns[:5])
                            if has_guid and 'код тн вэд' in ' '.join(normalized_columns[:5]).lower():
                                marketplace_type = 'vseinstrumenty'
                                confidence = 95.0
                                st.write("⚠️ DEBUG: Обнаружен шаблон Все инструменты по GUID* и коду ТН ВЭД")
                        
                        if marketplace_type != 'other' and confidence > 80:
                            st.session_state.source_marketplace = marketplace_type
                            # Выводим детали распознавания для отладки
                            st.markdown(f"<div style='font-size: 0.7rem; color: #aaa;'>DEBUG: {marketplace_type} (уверенность: {confidence:.1f}%)</div>", unsafe_allow_html=True)
                            
                            # Установка правильной строки заголовков по типу маркетплейса
                            if marketplace_type == 'ozon' and st.session_state.source_header_row != 2:
                                st.write(f"⚠️ DEBUG: Обнаружен Ozon, устанавливаем строку заголовков на 2")
                                st.session_state.source_header_row = 2
                                if "ozon_header_adjusted" not in st.session_state:
                                    st.session_state.ozon_header_adjusted = True
                                    st.rerun()
                            elif marketplace_type == 'wildberries' and st.session_state.source_header_row != 3:
                                st.write(f"⚠️ DEBUG: Обнаружен Wildberries, устанавливаем строку заголовков на 3")
                                st.session_state.source_header_row = 3
                                if "wb_header_adjusted" not in st.session_state:
                                    st.session_state.wb_header_adjusted = True
                                    st.rerun()
                            elif marketplace_type == 'lemanpro' and st.session_state.source_header_row != 4:
                                st.write(f"⚠️ DEBUG: Обнаружен ЛеманПро, устанавливаем строку заголовков на 4")
                                st.session_state.source_header_row = 4
                                if "lemanpro_header_adjusted" not in st.session_state:
                                    st.session_state.lemanpro_header_adjusted = True
                                    st.rerun()
                            # Если у нас Яндекс.Маркет, нужна специальная обработка для определения, в какой строке заголовки (2 или 4)
                            elif marketplace_type == 'yandex':
                                # Проверим, есть ли заголовки в 4-й строке для Яндекс.Маркет
                                sheet = source_workbook[selected_sheet]
                                yandex_header_row = 2  # По умолчанию строка 2
                                
                                if sheet.max_row >= 4:
                                    row_values_4 = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[4]]
                                    if any('ваш sku' in val for val in row_values_4) or any('качество карточки' in val for val in row_values_4):
                                        yandex_header_row = 4
                                
                                # Проверяем, нужно ли корректировать строку заголовков
                                if st.session_state.source_header_row != yandex_header_row:
                                    st.write(f"⚠️ DEBUG: Обнаружен Яндекс.Маркет с заголовками в строке {yandex_header_row}, корректируем")
                                    st.session_state.source_header_row = yandex_header_row
                                    if "yandex_header_adjusted" not in st.session_state:
                                        st.session_state.yandex_header_adjusted = True
                                        st.rerun()
            else:
                st.error("В исходном файле не найдено листов!")
                st.session_state.source_data = None
        except Exception as e:
            st.error(f"Ошибка при загрузке исходного файла: {str(e)}")
            st.session_state.source_data = None
    
    if st.session_state.source_workbook is not None and st.session_state.source_sheets:
        col1a, col1b = st.columns([3, 1])
        with col1a:
            selected_source_sheet = st.selectbox(
                "Выберите лист в исходной таблице", 
                st.session_state.source_sheets,
                index=st.session_state.source_sheets.index(st.session_state.source_sheet_name) if st.session_state.source_sheet_name in st.session_state.source_sheets else 0
            )
        
        with col1b:
            # Добавляем выбор строки с заголовками
            source_header_row = st.number_input(
                "Строка заголовков",
                min_value=1,
                max_value=20,
                value=st.session_state.source_header_row,
                step=1,
                key="source_header_input"
            )
        
        if selected_source_sheet != st.session_state.source_sheet_name or source_header_row != st.session_state.source_header_row:
            st.session_state.source_sheet_name = selected_source_sheet
            st.session_state.source_header_row = source_header_row
            st.session_state.auto_mapped = False
            st.session_state.mapping_complete = False
            st.session_state.transfer_complete = False
        
        try:
            sheet = st.session_state.source_workbook[st.session_state.source_sheet_name]
            # Определяем заголовки колонок (используем выбранную пользователем строку)
            header_row = st.session_state.source_header_row
            headers = []
            column_indices = []
            
            # Собираем заголовки и их индексы
            for i, cell in enumerate(sheet[header_row]):
                if cell.value is not None and str(cell.value).strip() != "":
                    headers.append(str(cell.value))
                    column_indices.append(i)
                
            # Проверяем на дубликаты и исправляем
            unique_headers = {}
            for i, header in enumerate(headers):
                if header in unique_headers:
                    # Если заголовок уже существует, добавляем суффикс
                    counter = 1
                    new_header = f"{header}_{counter}"
                    while new_header in unique_headers:
                        counter += 1
                        new_header = f"{header}_{counter}"
                    headers[i] = new_header
                unique_headers[headers[i]] = True
                
            # Читаем данные начиная со следующей строки после заголовка
            data = []
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if any(cell is not None for cell in row):
                    # Берем только данные из столбцов с заголовками
                    row_data = [row[idx] for idx in column_indices]
                    data.append(row_data)
            
            # Создаем DataFrame только с непустыми заголовками
            df = pd.DataFrame(data, columns=headers)
            
            # Преобразуем все данные в строки для избежания ошибок конвертации
            df = df.astype(str)
            
            st.session_state.source_data = df
            st.session_state.source_columns = headers
            
            # Показываем предпросмотр исходной таблицы
            st.write("Предпросмотр исходной таблицы:")
            st.dataframe(df.head(5))
            
            # Определяем и отображаем маркетплейс
            if st.session_state.source_columns:
                # Проверяем на ЛеманПро по содержимому
                lemanpro_detected = False
                
                # Особая логика для первой строки: проверка на наличие GUID в первых строках данных
                if st.session_state.source_header_row == 1:
                    # Отладочное сообщение
                    st.write(f"⚠️ DEBUG: Проверяем строку 1 на ЛеманПро")
                    
                    # Проверяем, есть ли в данных строки с текстом "GUID"
                    if st.session_state.source_data is not None and not st.session_state.source_data.empty:
                        first_rows = st.session_state.source_data.head(5).astype(str)
                        # Отладочное сообщение - просмотр первых строк
                        st.write(f"⚠️ DEBUG: Первые строки данных: {first_rows.values.tolist()}")
                        
                        for _, row in first_rows.iterrows():
                            row_text = " ".join(row.values).lower()
                            # Отладочное сообщение
                            st.write(f"⚠️ DEBUG: Проверяем текст: {row_text[:50]}...")
                            
                            if "guid" in row_text or "идентификатор из 1с" in row_text:
                                lemanpro_detected = True
                                marketplace = "lemanpro"
                                confidence = 95.0
                                st.write(f"⚠️ DEBUG: Обнаружен ЛеманПро по GUID!")
                                # Устанавливаем строку заголовков на 4 и делаем rerun для обновления UI
                                st.session_state.source_header_row = 4
                                st.write(f"⚠️ DEBUG: Устанавливаем строку заголовков ЛеманПро на 4")
                                # Добавляем флаг для предотвращения бесконечного цикла
                                if "lemanpro_header_adjusted" not in st.session_state:
                                    st.session_state.lemanpro_header_adjusted = True
                                    st.rerun()
                                break
                                
                # Если строка заголовка равна 4, это может быть ЛеманПро или Яндекс.Маркет
                elif st.session_state.source_header_row == 4:
                    # Убираем отладочные сообщения
                    
                    # Эталонные заголовки Яндекс.Маркет
                    yandex_reference_headers = [
                        'ваш sku *', 'качество карточки', 'рекомендации по заполнению',
                        'название товара *', 'ссылка на изображение', 'название группы вариантов', 
                        'штрихкод *', 'тип уценки', 'внешний вид товара', 'описание состояния товара',
                        'вес, кг'
                    ]
                    
                    # Эталонные заголовки ЛеманПро
                    lemanpro_reference_headers = [
                        'guid', 'код тн вэд', 'наименование товара мерчанта', 
                        'бренд товара', 'изготовитель', 'вес упаковки, кг', 
                        'габариты упаковки, см'
                    ]
                    
                    # Нормализуем колонки для сравнения
                    normalized_columns = [str(col).lower() for col in st.session_state.source_columns]
                    
                    # Подсчитываем количество совпадений с эталонными заголовками
                    yandex_matches = sum(1 for header in yandex_reference_headers if any(header.lower() in col for col in normalized_columns))
                    lemanpro_matches = sum(1 for header in lemanpro_reference_headers if any(header.lower() in col for col in normalized_columns))
                    
                    # Определяем тип маркетплейса на основе количества совпадений
                    if lemanpro_matches >= 2 or any('guid' in col for col in normalized_columns):
                        lemanpro_detected = True
                        marketplace = "lemanpro"
                        confidence = 95.0
                    elif yandex_matches >= 2:
                        marketplace = "yandex"
                        confidence = 95.0
                # Если строка заголовка равна 2, это может быть Ozon или Яндекс.Маркет
                elif st.session_state.source_header_row == 2:
                    # Проверяем наличие характерных заголовков Ozon и Яндекс.Маркета
                    ozon_indicators = ['название товара*', 'ссылка на главное фото*', 'артикул*', 'бренд*', 'ндс, %*', 'цена, руб.*', 'обязательное поле']
                    yandex_indicators = ['ваш sku *', 'качество карточки', 'фид', 'товар', 'цена']
                    
                    normalized_columns = [str(col).lower() for col in st.session_state.source_columns]
                    normalized_text = ' '.join(normalized_columns)
                    
                    # Проверяем название листа - Ozon часто использует именно "Шаблон"
                    is_ozon_sheet = False
                    if hasattr(st.session_state, 'source_sheet_name') and "шаблон" in str(st.session_state.source_sheet_name).lower():
                        is_ozon_sheet = True
                    
                    # Проверяем название листа для Яндекс.Маркет
                    is_yandex_sheet = False
                    if hasattr(st.session_state, 'source_sheet_name'):
                        is_yandex_sheet = "данные о товар" in str(st.session_state.source_sheet_name).lower()
                    
                    # ПРИНУДИТЕЛЬНАЯ ПРОВЕРКА НА OZON ПО ОБЯЗАТЕЛЬНЫМ ПРИЗНАКАМ
                    is_definitely_ozon = (
                        any('обязательное поле' in col for col in normalized_columns) or
                        any('цена, руб*' in col for col in normalized_columns) or
                        any('название товара*' in col for col in normalized_columns) or
                        any('ссылка на главное фото*' in col for col in normalized_columns) or
                        any('ссылки на дополнительные фото' in col for col in normalized_columns)
                    )
                    
                    # Проверка наличия звездочек (специфично для Ozon)
                    asterisk_count = sum(1 for col in normalized_columns if '*' in col)
                    
                    # Счетчики для каждого маркетплейса
                    found_ozon = 0
                    found_yandex = 0
                    
                    for col in normalized_columns:
                        # Проверка Ozon
                        for indicator in ozon_indicators:
                            if indicator in col:
                                found_ozon += 1
                                break
                                
                        # Проверка Яндекс.Маркет
                        for indicator in yandex_indicators:
                            if indicator in col:
                                found_yandex += 1
                                break
                    
                    # Особые признаки Ozon
                    has_ozon_links = any('ссылка на' in col for col in normalized_columns)
                    has_ozon_price = any('цена, руб' in col for col in normalized_columns)
                    
                    # Особые признаки Яндекс.Маркет
                    has_yandex_sku = any('ваш sku' in col for col in normalized_columns)
                    has_quality = any('качество карточки' in col for col in normalized_columns)
                    
                    # Проверка на конкретные признаки Яндекс.Маркет
                    is_yandex_pattern = (
                        ("ваш sku" in normalized_text and "качество карточки" in normalized_text) or
                        ("ваш sku *" in normalized_text) or
                        (any('param_ids' in col for col in normalized_columns) and any('param_names' in col for col in normalized_columns))
                    )
                    
                    # Определение маркетплейса на основе найденных признаков
                    if is_definitely_ozon or (found_ozon >= 2) or (is_ozon_sheet and found_ozon >= 1) or (asterisk_count >= 3 and has_ozon_links):
                        marketplace = "ozon"
                        confidence = 95.0
                    elif is_yandex_pattern or (found_yandex >= 1) or (is_yandex_sheet) or has_yandex_sku or has_quality:
                        marketplace = "yandex"
                        confidence = 95.0
                    else:
                        # Если не смогли определить по признакам, используем стандартную функцию
                        marketplace, confidence = detect_marketplace_template(st.session_state.source_columns)
                else:
                    marketplace, confidence = detect_marketplace_template(st.session_state.source_columns)
                    
                if marketplace != 'other':
                    if marketplace == 'wildberries':
                        mp_name = "Wildberries"
                        mp_color = "purple"
                        marketplace_icon = "attached_assets/wildberries.png"
                    elif marketplace == 'ozon':
                        mp_name = "Ozon"
                        mp_color = "blue"
                        marketplace_icon = "attached_assets/ozon.png"
                    elif marketplace == 'lemanpro':
                        mp_name = "ЛеманПро"
                        mp_color = "green"
                        marketplace_icon = "attached_assets/Лемана про.png"
                    elif marketplace == 'yandex':
                        mp_name = "Яндекс.Маркет"
                        mp_color = "orange"
                        marketplace_icon = "attached_assets/Яндекс маркет.png"
                    else:
                        mp_name = marketplace.capitalize()
                        mp_color = "gray"
                        marketplace_icon = "attached_assets/xlsx.png"
                        # Проверяем другие возможные маркетплейсы
                        if "сбер" in marketplace.lower():
                            marketplace_icon = "attached_assets/сбермегамаркет.png"
                        elif "инструмент" in marketplace.lower():
                            marketplace_icon = "attached_assets/все инструменты.png"
                    
                    st.caption(f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64(marketplace_icon)}" width="20" style="margin-right:5px"></span> Распознан шаблон: <span style="color:{mp_color};font-weight:bold;">{mp_name}</span> (уверенность: {confidence:.1f}%)', unsafe_allow_html=True)
            
        except Exception as e:
            st.error(f"Ошибка при обработке исходного файла: {str(e)}")
            st.session_state.source_data = None

with col2:
    st.subheader("📥 Целевая таблица (Куда)")
    
    # Добавляем логотипы маркетплейсов в области загрузки файлов в один ряд маленькими иконками
    st.write(
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/xlsx.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/ozon.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/wildberries.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/Яндекс маркет.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/сбермегамаркет.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/все инструменты.png")}" width="18" style="margin-right:4px"></span>'
        f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64("attached_assets/Лемана про.png")}" width="18" style="margin-right:4px"></span>',
        unsafe_allow_html=True
    )
    
    target_file = st.file_uploader("Загрузите целевую таблицу (xlsx)", type=['xlsx'], key="target_uploader")
    
    if target_file is not None and target_file != st.session_state.target_file:
        st.session_state.target_file = target_file
        try:
            target_workbook, target_sheets = load_excel_file(target_file)
            st.session_state.target_workbook = target_workbook
            st.session_state.target_sheets = target_sheets
            
            if len(target_sheets) > 0:
                # По умолчанию используем первый лист
                selected_sheet = target_sheets[0]
                header_row = 1
                
                # Маркетплейсы и характерные для них листы и строки заголовков
                marketplace_config = {
                    'ozon': {
                        'sheet_patterns': ['шаблон', 'template', 'озон', 'ozon'],
                        'header_row': 2
                    },
                    'wildberries': {
                        'sheet_patterns': ['товары', 'вб', 'wb', 'wildberries'],
                        'header_row': 3
                    },
                    'lemanpro': {
                        'sheet_patterns': ['леман', 'атем', 'leman', 'atem'],
                        'header_row': 4
                    },
                    'yandex': {
                        'sheet_patterns': ['данные о товарах', 'яндекс', 'маркет', 'яндекс маркет', 'yandex'],
                        'header_row': 2
                    }
                }
                
                # Шаг 1: Ищем подходящий лист по имени
                sheet_found = False
                
                # Специальная проверка для ЛеманПро по имени файла
                target_filename = getattr(target_file, 'name', '')
                if target_filename:
                    # Для ЛеманПро проверяем соответствие имени файла и названия листа
                    file_base_name = os.path.splitext(os.path.basename(target_filename))[0]
                    
                    for sheet_name in target_sheets:
                        # Ищем лист, в названии которого есть "шаблон" и имя файла
                        # или характерные для ЛеманПро имена
                        if ("шаблон" in sheet_name.lower() and file_base_name.lower() in sheet_name.lower()) or \
                           any(pattern in sheet_name.lower() for pattern in ["атем", "atem", "леман"]):
                            selected_sheet = sheet_name
                            header_row = 4  # Для ЛеманПро характерна 4-я строка заголовков
                            marketplace_type = 'lemanpro'
                            sheet_found = True
                            break
                
                # Если не нашли ЛеманПро, ищем другие маркетплейсы
                if not sheet_found:
                    for sheet_name in target_sheets:
                        sheet_lower = sheet_name.lower()
                        for marketplace, config in marketplace_config.items():
                            if any(pattern in sheet_lower for pattern in config['sheet_patterns']):
                                selected_sheet = sheet_name
                                header_row = config['header_row']
                                marketplace_type = marketplace
                                sheet_found = True
                                break
                        if sheet_found:
                            break
                
                # Устанавливаем лист и строку заголовков
                st.session_state.target_sheet_name = selected_sheet
                st.session_state.target_header_row = header_row
                
                # Шаг 2: Получаем заголовки из выбранного листа
                sheet = target_workbook[selected_sheet]
                if sheet.max_row >= header_row:
                    # Собираем заголовки
                    headers = []
                    for cell in sheet[header_row]:
                        if cell.value is not None and str(cell.value).strip() != "":
                            headers.append(str(cell.value))
                    
                    # Шаг 3: Определяем маркетплейс по заголовкам с учетом строки и первых 5 колонок
                    if headers:
                        # Используем новый модуль распознавания для более точного определения
                        normalized_columns = [str(h).lower().strip() for h in headers]
                        marketplace_type, confidence, details = marketplace_detection.detect_marketplace_by_row_headers(
                            normalized_columns, 
                            st.session_state.target_header_row
                        )
                        
                        # Проверяем, есть ли в названии листа "Данные" для "Все инструменты"
                        sheet_name_lower = selected_sheet.lower()
                        if 'данные' in sheet_name_lower and st.session_state.target_header_row == 2:
                            # Проверяем на шаблон Все инструменты
                            has_guid = any('guid*' in col.lower() for col in normalized_columns[:5])
                            if has_guid and 'код тн вэд' in ' '.join(normalized_columns[:5]).lower():
                                marketplace_type = 'vseinstrumenty'
                                confidence = 95.0
                                st.write("⚠️ DEBUG: Обнаружен шаблон Все инструменты (целевой) по GUID* и коду ТН ВЭД")
                        
                        if marketplace_type != 'other' and confidence > 80:
                            st.session_state.target_marketplace = marketplace_type
                            # Выводим детали распознавания для отладки
                            st.markdown(f"<div style='font-size: 0.7rem; color: #aaa;'>DEBUG: {marketplace_type} (уверенность: {confidence:.1f}%)</div>", unsafe_allow_html=True)
                            
                            # Установка правильной строки заголовков по типу маркетплейса для целевого файла
                            if marketplace_type == 'ozon' and st.session_state.target_header_row != 2:
                                st.write(f"⚠️ DEBUG: Обнаружен Ozon, устанавливаем строку заголовков целевого файла на 2")
                                st.session_state.target_header_row = 2
                                if "target_ozon_header_adjusted" not in st.session_state:
                                    st.session_state.target_ozon_header_adjusted = True
                                    st.rerun()
                            # Если у нас Яндекс.Маркет, нужна специальная обработка для определения, в какой строке заголовки (2 или 4)
                            elif marketplace_type == 'yandex':
                                # Проверим, есть ли заголовки в 4-й строке для Яндекс.Маркет
                                sheet = target_workbook[selected_sheet]
                                yandex_header_row = 2  # По умолчанию строка 2
                                
                                if sheet.max_row >= 4:
                                    row_values_4 = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[4]]
                                    if any('ваш sku' in val for val in row_values_4) or any('качество карточки' in val for val in row_values_4):
                                        yandex_header_row = 4
                                
                                # Проверяем, нужно ли корректировать строку заголовков
                                if st.session_state.target_header_row != yandex_header_row:
                                    st.write(f"⚠️ DEBUG: Обнаружен Яндекс.Маркет с заголовками в строке {yandex_header_row}, корректируем")
                                    st.session_state.target_header_row = yandex_header_row
                                    if "target_yandex_header_adjusted" not in st.session_state:
                                        st.session_state.target_yandex_header_adjusted = True
                                        st.rerun()
                            elif marketplace_type == 'wildberries' and st.session_state.target_header_row != 3:
                                st.write(f"⚠️ DEBUG: Обнаружен Wildberries, устанавливаем строку заголовков целевого файла на 3")
                                st.session_state.target_header_row = 3
                                if "target_wb_header_adjusted" not in st.session_state:
                                    st.session_state.target_wb_header_adjusted = True
                                    st.rerun()
                            elif marketplace_type == 'lemanpro' and st.session_state.target_header_row != 4:
                                st.write(f"⚠️ DEBUG: Обнаружен ЛеманПро, устанавливаем строку заголовков целевого файла на 4")
                                st.session_state.target_header_row = 4
                                if "target_lemanpro_header_adjusted" not in st.session_state:
                                    st.session_state.target_lemanpro_header_adjusted = True
                                    st.rerun()
            else:
                st.error("В целевом файле не найдено листов!")
                st.session_state.target_data = None
        except Exception as e:
            st.error(f"Ошибка при загрузке целевого файла: {str(e)}")
            st.session_state.target_data = None
    
    if st.session_state.target_workbook is not None and st.session_state.target_sheets:
        col2a, col2b = st.columns([3, 1])
        with col2a:
            selected_target_sheet = st.selectbox(
                "Выберите лист в целевой таблице", 
                st.session_state.target_sheets,
                index=st.session_state.target_sheets.index(st.session_state.target_sheet_name) if st.session_state.target_sheet_name in st.session_state.target_sheets else 0
            )
        
        with col2b:
            # Добавляем выбор строки с заголовками
            target_header_row = st.number_input(
                "Строка заголовков",
                min_value=1,
                max_value=20,
                value=st.session_state.target_header_row,
                step=1,
                key="target_header_input"
            )
        
        if selected_target_sheet != st.session_state.target_sheet_name or target_header_row != st.session_state.target_header_row:
            st.session_state.target_sheet_name = selected_target_sheet
            st.session_state.target_header_row = target_header_row
            st.session_state.auto_mapped = False
            st.session_state.mapping_complete = False
            st.session_state.transfer_complete = False
        
        try:
            sheet = st.session_state.target_workbook[st.session_state.target_sheet_name]
            # Определяем заголовки колонок (используем выбранную пользователем строку)
            header_row = st.session_state.target_header_row
            headers = []
            column_indices = []
            
            # Собираем заголовки и их индексы
            for i, cell in enumerate(sheet[header_row]):
                if cell.value is not None and str(cell.value).strip() != "":
                    headers.append(str(cell.value))
                    column_indices.append(i)
                
            # Проверяем на дубликаты и исправляем
            unique_headers = {}
            for i, header in enumerate(headers):
                if header in unique_headers:
                    # Если заголовок уже существует, добавляем суффикс
                    counter = 1
                    new_header = f"{header}_{counter}"
                    while new_header in unique_headers:
                        counter += 1
                        new_header = f"{header}_{counter}"
                    headers[i] = new_header
                unique_headers[headers[i]] = True
                
            # Читаем данные начиная со следующей строки после заголовка
            data = []
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if any(cell is not None for cell in row):
                    # Берем только данные из столбцов с заголовками
                    row_data = [row[idx] for idx in column_indices]
                    data.append(row_data)
            
            # Создаем DataFrame только с непустыми заголовками
            df = pd.DataFrame(data, columns=headers)
            
            # Преобразуем все данные в строки для избежания ошибок конвертации
            df = df.astype(str)
            
            st.session_state.target_data = df
            st.session_state.target_columns = headers
            
            # Показываем предпросмотр целевой таблицы
            st.write("Предпросмотр целевой таблицы:")
            st.dataframe(df.head(5))
            
            # Определяем и отображаем маркетплейс
            if st.session_state.target_columns:
                # Проверяем на ЛеманПро по содержимому
                lemanpro_detected = False
                
                # Особая логика для первой строки: проверка на наличие GUID в первых строках данных
                if st.session_state.target_header_row == 1:
                    # Проверяем, есть ли в данных строки с текстом "GUID"
                    if st.session_state.target_data is not None and not st.session_state.target_data.empty:
                        first_rows = st.session_state.target_data.head(5).astype(str)
                        for _, row in first_rows.iterrows():
                            row_text = " ".join(row.values).lower()
                            if "guid" in row_text or "идентификатор из 1с" in row_text:
                                lemanpro_detected = True
                                marketplace = "lemanpro"
                                confidence = 95.0
                                # Устанавливаем строку заголовков на 4 и делаем rerun для обновления UI
                                st.session_state.target_header_row = 4
                                # Добавляем флаг для предотвращения бесконечного цикла
                                if "target_lemanpro_header_adjusted" not in st.session_state:
                                    st.session_state.target_lemanpro_header_adjusted = True
                                    st.rerun()
                                break
                                
                # Если строка заголовка равна 4, это может быть ЛеманПро или Яндекс.Маркет
                elif st.session_state.target_header_row == 4:
                    normalized_columns = [str(col).lower() for col in st.session_state.target_columns]
                    
                    # Эталонные заголовки Яндекс.Маркет
                    yandex_reference_headers = [
                        'ваш sku *', 'качество карточки', 'рекомендации по заполнению',
                        'название товара *', 'ссылка на изображение', 'название группы вариантов', 
                        'штрихкод *', 'тип уценки', 'внешний вид товара', 'описание состояния товара',
                        'вес, кг'
                    ]
                    
                    # Эталонные заголовки ЛеманПро
                    lemanpro_reference_headers = [
                        'guid', 'код тн вэд', 'наименование товара мерчанта', 
                        'бренд товара', 'изготовитель', 'вес упаковки, кг', 
                        'габариты упаковки, см'
                    ]
                    
                    # Подсчитываем количество совпадений с эталонными заголовками
                    yandex_matches = sum(1 for header in yandex_reference_headers if any(header.lower() in col for col in normalized_columns))
                    lemanpro_matches = sum(1 for header in lemanpro_reference_headers if any(header.lower() in col for col in normalized_columns))
                    
                    # Определяем тип маркетплейса на основе количества совпадений
                    if lemanpro_matches >= 2 or any('guid' in col for col in normalized_columns):
                        lemanpro_detected = True
                        marketplace = "lemanpro"
                        confidence = 95.0
                    elif yandex_matches >= 2:
                        marketplace = "yandex"
                        confidence = 95.0
                
                # Если строка заголовка равна 2, это может быть Ozon или Яндекс.Маркет
                elif st.session_state.target_header_row == 2:
                    # Проверяем наличие характерных заголовков Ozon и Яндекс.Маркета
                    ozon_indicators = ['название товара*', 'ссылка на главное фото*', 'артикул*', 'бренд*', 'ндс, %*', 'цена, руб.*', 'обязательное поле']
                    yandex_indicators = ['ваш sku *', 'качество карточки', 'фид', 'товар', 'цена']
                    
                    normalized_columns = [str(col).lower() for col in st.session_state.target_columns]
                    normalized_text = ' '.join(normalized_columns)
                    
                    # Проверяем название листа - Ozon часто использует именно "Шаблон"
                    is_ozon_sheet = False
                    if hasattr(st.session_state, 'target_sheet_name') and "шаблон" in str(st.session_state.target_sheet_name).lower():
                        is_ozon_sheet = True
                    
                    # Проверяем название листа для Яндекс.Маркет
                    is_yandex_sheet = False
                    if hasattr(st.session_state, 'target_sheet_name'):
                        is_yandex_sheet = "данные о товар" in str(st.session_state.target_sheet_name).lower()
                    
                    # ПРИНУДИТЕЛЬНАЯ ПРОВЕРКА НА OZON ПО ОБЯЗАТЕЛЬНЫМ ПРИЗНАКАМ
                    is_definitely_ozon = (
                        any('обязательное поле' in col for col in normalized_columns) or
                        any('цена, руб*' in col for col in normalized_columns) or
                        any('название товара*' in col for col in normalized_columns) or
                        any('ссылка на главное фото*' in col for col in normalized_columns) or
                        any('ссылки на дополнительные фото' in col for col in normalized_columns)
                    )
                    
                    # Проверка наличия звездочек (специфично для Ozon)
                    asterisk_count = sum(1 for col in normalized_columns if '*' in col)
                    
                    # Счетчики для каждого маркетплейса
                    found_ozon = 0
                    found_yandex = 0
                    
                    for col in normalized_columns:
                        # Проверка Ozon
                        for indicator in ozon_indicators:
                            if indicator in col:
                                found_ozon += 1
                                break
                                
                        # Проверка Яндекс.Маркет
                        for indicator in yandex_indicators:
                            if indicator in col:
                                found_yandex += 1
                                break
                    
                    # Особые признаки Ozon
                    has_ozon_links = any('ссылка на' in col for col in normalized_columns)
                    has_ozon_price = any('цена, руб' in col for col in normalized_columns)
                    
                    # Особые признаки Яндекс.Маркет
                    has_yandex_sku = any('ваш sku' in col for col in normalized_columns)
                    has_quality = any('качество карточки' in col for col in normalized_columns)
                    
                    # Проверка на конкретные признаки Яндекс.Маркет
                    is_yandex_pattern = (
                        ("ваш sku" in normalized_text and "качество карточки" in normalized_text) or
                        ("ваш sku *" in normalized_text) or
                        (any('param_ids' in col for col in normalized_columns) and any('param_names' in col for col in normalized_columns))
                    )
                    
                    # Определение маркетплейса на основе найденных признаков
                    if is_definitely_ozon or (found_ozon >= 2) or (is_ozon_sheet and found_ozon >= 1) or (asterisk_count >= 3 and has_ozon_links):
                        marketplace = "ozon"
                        confidence = 95.0
                    elif is_yandex_pattern or (found_yandex >= 1) or (is_yandex_sheet) or has_yandex_sku or has_quality:
                        marketplace = "yandex"
                        confidence = 95.0
                    else:
                        # Если не смогли определить по признакам, используем стандартную функцию
                        marketplace, confidence = detect_marketplace_template(st.session_state.target_columns)
                
                # Если строка заголовка равна 3, вероятно это Wildberries
                elif st.session_state.target_header_row == 3:
                    # Проверяем наличие характерных для WB заголовков
                    wb_indicators = ['артикул wb', 'баркод', 'номенклатура', 'номер номенклатуры', 'ставка ндс', 'предмет']
                    normalized_columns = [str(col).lower() for col in st.session_state.target_columns]
                    
                    # Проверяем название листа для WB
                    is_wb_sheet = False
                    if hasattr(st.session_state, 'target_sheet_name'):
                        is_wb_sheet = "товары" in str(st.session_state.target_sheet_name).lower()
                    
                    # Счетчик для WB
                    found_wb = 0
                    for col in normalized_columns:
                        for indicator in wb_indicators:
                            if indicator in col:
                                found_wb += 1
                                break
                    
                    # Уникальные признаки WB
                    has_wb_features = any('артикул wb' in col for col in normalized_columns) or any('номенклатур' in col for col in normalized_columns)
                    
                    # Определяем маркетплейс
                    if is_wb_sheet or found_wb >= 1 or has_wb_features:
                        marketplace = "wildberries"
                        confidence = 95.0
                    else:
                        marketplace, confidence = detect_marketplace_template(st.session_state.target_columns)
                
                # Если не было особых условий, используем стандартную функцию
                else:
                    marketplace, confidence = detect_marketplace_template(st.session_state.target_columns)
                
                if marketplace != 'other':
                    if marketplace == 'wildberries':
                        mp_name = "Wildberries"
                        mp_color = "purple"
                        marketplace_icon = "attached_assets/wildberries.png"
                    elif marketplace == 'ozon':
                        mp_name = "Ozon"
                        mp_color = "blue"
                        marketplace_icon = "attached_assets/ozon.png"
                    elif marketplace == 'lemanpro':
                        mp_name = "ЛеманПро"
                        mp_color = "green"
                        marketplace_icon = "attached_assets/Лемана про.png"
                    elif marketplace == 'yandex':
                        mp_name = "Яндекс.Маркет"
                        mp_color = "orange"
                        marketplace_icon = "attached_assets/Яндекс маркет.png"
                    else:
                        mp_name = marketplace.capitalize()
                        mp_color = "gray"
                        marketplace_icon = "attached_assets/xlsx.png"
                        # Проверяем другие возможные маркетплейсы
                        if "сбер" in marketplace.lower():
                            marketplace_icon = "attached_assets/сбермегамаркет.png"
                        elif "инструмент" in marketplace.lower():
                            marketplace_icon = "attached_assets/все инструменты.png"
                    
                    st.caption(f'<span style="display: inline-block;"><img src="data:image/png;base64,{get_image_base64(marketplace_icon)}" width="20" style="margin-right:5px"></span> Распознан шаблон: <span style="color:{mp_color};font-weight:bold;">{mp_name}</span> (уверенность: {confidence:.1f}%)', unsafe_allow_html=True)
            
        except Exception as e:
            st.error(f"Ошибка при обработке целевого файла: {str(e)}")
            st.session_state.target_data = None

st.divider()

# Раздел автоматического и ручного маппинга
if st.session_state.source_data is not None and st.session_state.target_data is not None:
    st.header("🔄 Сопоставление колонок")
    
    # Кнопка для автоматического маппинга
    if not st.session_state.auto_mapped:
        # Определяем иконки для кнопки маппинга на основе распознанных маркетплейсов
        source_marketplace = "other"
        target_marketplace = "other"
        
        if hasattr(st.session_state, 'source_columns'):
            # Используем нашу новую логику определения маркетплейса через функцию detect_marketplace_template
            source_marketplace, confidence = detect_marketplace_template(st.session_state.source_columns)
            
            # Если маркетплейс не определен или уверенность низкая, пробуем дополнительные проверки
            if source_marketplace == "other" or confidence < 80:
                sheet_name = st.session_state.source_sheet_name if hasattr(st.session_state, 'source_sheet_name') else ""
                
                # Проверка имени листа
                sheet_name_lower = sheet_name.lower()
                if "шаблон" in sheet_name_lower:
                    source_marketplace = "ozon"
                elif "товары" in sheet_name_lower:
                    source_marketplace = "wildberries"
                elif "данные о товарах" in sheet_name_lower:
                    source_marketplace = "yandex"
                elif "леман" in sheet_name_lower or "atem" in sheet_name_lower or "атем" in sheet_name_lower:
                    source_marketplace = "lemanpro"
        
        if hasattr(st.session_state, 'target_columns'):
            # Используем унифицированный подход определения маркетплейса через функцию detect_marketplace_template
            target_marketplace, confidence = detect_marketplace_template(st.session_state.target_columns)
            
            # Если маркетплейс не определен или уверенность низкая, пробуем дополнительные проверки
            if target_marketplace == "other" or confidence < 80:
                sheet_name = st.session_state.target_sheet_name if hasattr(st.session_state, 'target_sheet_name') else ""
                
                # Проверка имени листа
                sheet_name_lower = sheet_name.lower()
                if "шаблон" in sheet_name_lower:
                    target_marketplace = "ozon"
                elif "товары" in sheet_name_lower:
                    target_marketplace = "wildberries"
                elif "данные о товарах" in sheet_name_lower:
                    target_marketplace = "yandex"
                elif "леман" in sheet_name_lower or "atem" in sheet_name_lower or "атем" in sheet_name_lower:
                    target_marketplace = "lemanpro"
        
        mapping_icon_source = "attached_assets/xlsx.png"
        mapping_icon_target = "attached_assets/xlsx.png"
        
        # Выбираем иконки для источника и цели
        if source_marketplace == 'wildberries':
            mapping_icon_source = "attached_assets/wildberries.png"
        elif source_marketplace == 'ozon':
            mapping_icon_source = "attached_assets/ozon.png"
        elif source_marketplace == 'lemanpro':
            mapping_icon_source = "attached_assets/Лемана про.png"
        elif "яндекс" in source_marketplace.lower():
            mapping_icon_source = "attached_assets/Яндекс маркет.png"
        
        if target_marketplace == 'wildberries':
            mapping_icon_target = "attached_assets/wildberries.png"
        elif target_marketplace == 'ozon':
            mapping_icon_target = "attached_assets/ozon.png"
        elif target_marketplace == 'lemanpro':
            mapping_icon_target = "attached_assets/Лемана про.png"
        elif "яндекс" in target_marketplace.lower():
            mapping_icon_target = "attached_assets/Яндекс маркет.png"
            
        # HTML не поддерживается в кнопках, используем текстовые обозначения
        if st.button(f"🔄 Автоматический маппинг колонок", use_container_width=True):
            with st.spinner("Выполняется автоматическое сопоставление колонок..."):
                st.session_state.column_mapping = map_columns_automatically(
                    st.session_state.source_columns,
                    st.session_state.target_columns
                )
                st.session_state.auto_mapped = True
                st.rerun()
    
    # Отображение и редактирование маппинга
    if st.session_state.auto_mapped:
        st.subheader("Настройка сопоставления колонок")
        
        # Получаем список всех колонок целевой таблицы для выбора
        all_target_columns = ["Не переносить"] + list(st.session_state.target_columns)
        
        # Разделяем колонки на автоматически сопоставленные и несопоставленные
        auto_mapped_columns = []
        unmapped_columns = []
        
        for src_col in st.session_state.source_columns:
            if src_col in st.session_state.column_mapping:
                auto_mapped_columns.append(src_col)
            else:
                unmapped_columns.append(src_col)
        
        # Создаем форму для сопоставления
        with st.form(key="mapping_form"):
            st.markdown("### ✅ Автоматически сопоставленные колонки:")
            st.caption("Слева - исходные заголовки, справа - выбор целевых заголовков.")
            
            # Создаем маппинги для автоматически сопоставленных колонок
            for src_col in auto_mapped_columns:
                # Определяем текущее значение маппинга
                current_mapping = st.session_state.column_mapping.get(src_col, None)
                
                # Определяем индекс текущего выбора
                selected_idx = 0
                if current_mapping is not None:
                    try:
                        selected_idx = all_target_columns.index(current_mapping)
                    except ValueError:
                        # Если значение не найдено, добавляем его в список
                        all_target_columns.append(current_mapping)
                        selected_idx = len(all_target_columns) - 1
                
                # Компактное отображение колонок
                cols = st.columns([3, 3])
                with cols[0]:
                    st.markdown(f"**{src_col}**", help="Исходная колонка")
                with cols[1]:
                    mapping = st.selectbox(
                        "Целевая колонка", 
                        options=all_target_columns,
                        index=selected_idx,
                        key=f"auto_map_{src_col}",
                        label_visibility="collapsed"
                    )
                
                # Обновляем маппинг
                if mapping != "Не переносить":
                    st.session_state.column_mapping[src_col] = mapping
                elif src_col in st.session_state.column_mapping:
                    del st.session_state.column_mapping[src_col]
            
            # Добавляем отступ между группами
            st.divider()
            
            # Если есть несопоставленные колонки
            if unmapped_columns:
                st.markdown("### ❌ Несопоставленные колонки:")
                st.caption("Для этих колонок нужно вручную выбрать целевые заголовки из выпадающего списка.")
                
                # Создаем маппинги для несопоставленных колонок
                for src_col in unmapped_columns:
                    # Определяем текущее значение маппинга (должно быть None)
                    current_mapping = st.session_state.column_mapping.get(src_col, None)
                    
                    # Определяем индекс текущего выбора
                    selected_idx = 0
                    if current_mapping is not None:
                        try:
                            selected_idx = all_target_columns.index(current_mapping)
                        except ValueError:
                            all_target_columns.append(current_mapping)
                            selected_idx = len(all_target_columns) - 1
                    
                    # Компактное отображение колонок
                    cols = st.columns([3, 3])
                    with cols[0]:
                        st.markdown(f"**{src_col}**", help="Исходная колонка")
                    with cols[1]:
                        mapping = st.selectbox(
                            "Целевая колонка",
                            options=all_target_columns,
                            index=selected_idx,
                            key=f"unmap_{src_col}",
                            label_visibility="collapsed"
                        )
                    
                    # Обновляем маппинг
                    if mapping != "Не переносить":
                        st.session_state.column_mapping[src_col] = mapping
                    elif src_col in st.session_state.column_mapping:
                        del st.session_state.column_mapping[src_col]
            
            # Показываем сводку по сопоставлению
            total_src = len(st.session_state.source_columns)
            total_mapped = len(st.session_state.column_mapping)
            st.info(f"Сопоставлено {total_mapped} из {total_src} исходных колонок")
            
            # Кнопка для завершения маппинга
            submitted = st.form_submit_button("✅ Подтвердить сопоставление")
            if submitted:
                st.session_state.mapping_complete = True
                st.success("Сопоставление колонок выполнено успешно!")
                st.rerun()
    
    # Отображение результатов маппинга
    if st.session_state.mapping_complete:
        st.subheader("Результаты сопоставления")
        
        # Создаем таблицу сопоставлений
        mapping_data = []
        for src_col, tgt_col in st.session_state.column_mapping.items():
            mapping_data.append({"Исходная колонка": src_col, "Целевая колонка": tgt_col})
        
        if mapping_data:
            mapping_df = pd.DataFrame(mapping_data)
            st.dataframe(mapping_df, use_container_width=True)
        else:
            st.warning("Нет сопоставленных колонок!")
        
        # Кнопка для выполнения переноса данных
        if not st.session_state.transfer_complete:
            if st.button("📤 Выполнить перенос данных"):
                with st.spinner("Выполняется перенос данных..."):
                    try:
                        # Предварительный просмотр результата
                        preview_df = preview_data(
                            st.session_state.source_data, 
                            st.session_state.target_data,
                            st.session_state.column_mapping,
                            st.session_state.source_file.name if st.session_state.source_file else None
                        )
                        st.session_state.preview_result = preview_df
                        st.session_state.transfer_complete = True
                        st.rerun()
                    except Exception as e:
                        st.error(f"Ошибка при переносе данных: {str(e)}")
                        st.session_state.transfer_complete = False
        
        # Отображение результатов переноса и кнопка скачивания
        if st.session_state.transfer_complete and st.session_state.preview_result is not None:
            st.subheader("Предпросмотр результата")
            st.dataframe(st.session_state.preview_result.head(10), use_container_width=True)
            
            # Кнопка для скачивания результата
            if st.button("💾 Скачать обновленный файл"):
                with st.spinner("Подготовка файла для скачивания..."):
                    try:
                        # Подготовка обновленного файла
                        output = io.BytesIO()
                        
                        # Перенос данных в целевой файл с сохранением форматирования
                        result_workbook = transfer_data_between_tables(
                            st.session_state.source_data,
                            st.session_state.target_workbook,
                            st.session_state.target_sheet_name,
                            st.session_state.column_mapping,
                            st.session_state.target_header_row,
                            st.session_state.source_file.name if st.session_state.source_file else None
                        )
                        
                        # Сохранение результата в BytesIO буфер
                        result_workbook.save(output)
                        output.seek(0)
                        
                        # Определение имени выходного файла
                        original_filename = st.session_state.target_file.name
                        filename_parts = os.path.splitext(original_filename)
                        output_filename = f"{filename_parts[0]}_обновленный{filename_parts[1]}"
                        
                        # Скачивание файла
                        st.download_button(
                            label="📥 Скачать результат",
                            data=output,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.success(f"Файл '{output_filename}' готов к скачиванию!")
                        
                    except Exception as e:
                        st.error(f"Ошибка при подготовке файла: {str(e)}")
        
        # Кнопка для сброса маппинга и начала заново
        if st.button("🔄 Сбросить и начать заново"):
            st.session_state.column_mapping = {}
            st.session_state.mapping_complete = False
            st.session_state.transfer_complete = False
            st.session_state.auto_mapped = False
            st.session_state.preview_result = None
            st.rerun()

# Инструкции и пояснения
with st.expander("ℹ️ Инструкция по использованию"):
    st.markdown("""
    ### Как пользоваться инструментом:
    
    1. **Загрузка файлов**:
       - В левой колонке загрузите исходный файл, из которого будут взяты данные
       - В правой колонке загрузите целевой файл, в который будут переноситься данные
       - При необходимости выберите нужные листы в каждом файле
       - Укажите номер строки, содержащей заголовки для каждой таблицы
    
    2. **Сопоставление колонок**:
       - Нажмите кнопку "Автоматический маппинг колонок" для автоматического сопоставления
       - Система автоматически найдет соответствия между колонками на основе их названий
       - Проверьте результаты сопоставления в таблице
       - Нажмите "Подтвердить сопоставление" для завершения настройки
    
    3. **Перенос данных**:
       - После подтверждения сопоставления нажмите "Выполнить перенос данных"
       - Проверьте предпросмотр результата
       - Нажмите "Скачать обновленный файл" для получения файла с перенесенными данными
    
    4. **Начать заново**:
       - Для сброса текущего сопоставления и начала нового нажмите "Сбросить и начать заново"
    
    ### Особенности работы:
    - Инструмент сохраняет исходное форматирование и структуру целевого файла
    - Поддерживаются файлы Excel (.xlsx) от Ozon, Wildberries и других источников
    - Автоматический маппинг основан на сходстве заголовков колонок (например, "Артикул" и "Код товара")
    - Приложение игнорирует пустые столбцы и автоматически обрабатывает дублирующиеся заголовки
    """)

# Добавляем информацию в нижней части основного интерфейса
st.divider()
st.caption("""
**О приложении**: Инструмент для маппинга и переноса данных между таблицами маркетплейсов. 
Поддерживает работу с Ozon, Wildberries и другими Excel-шаблонами. **Версия:** 1.0
""")
